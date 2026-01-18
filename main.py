import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import csv
from datetime import datetime
import locale
from xml.sax.saxutils import escape
import os
import sys 
import shutil
import threading
import queue
import subprocess
import time
import urllib3

# --- Selenium Imports ---
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- SCRAPER CONFIGURATION ---
TARGET_URL = "https://ecatalogue.heromotocorp.biz:8080/Hero/"
TEMP_PROFILE_DIR = r"C:\selenium\HeroEdgeProfile"
gui_queue = queue.Queue()
active_driver = None
automation_process = None

# --- LOCATE DRIVER ---
# --- LOCATE DRIVER ---
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Not bundled, use the script's directory
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

# Correctly locate driver whether in PyCharm or as an EXE
DRIVER_PATH = resource_path("msedgedriver.exe")

def find_edge_executable():
    possible_paths = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\Application\msedge.exe")
    ]
    for path in possible_paths:
        if os.path.exists(path): return path
    return None
BROWSER_EXE = find_edge_executable()


# --- Calendar Widget (Requires 'pip install tkcalendar') ---
try:
    from tkcalendar import DateEntry
    CALENDAR_ENABLED = True
except ImportError:
    CALENDAR_ENABLED = False  # <--- THIS IS HAPPENING

# --- EXISTING IMPORTS (Calendar, ReportLab, etc) ---
try:
    from tkcalendar import DateEntry
    CALENDAR_ENABLED = True
except ImportError:
    CALENDAR_ENABLED = False

try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    PDF_ENABLED = True
except ImportError:
    PDF_ENABLED = False

# --- NEW: XLSX Generation (Requires 'pip install openpyxl') ---
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    XLSX_ENABLED = True
except ImportError:
    XLSX_ENABLED = False
# --- END NEW ---

# Set locale for Indian currency formatting
try:
    locale.setlocale(locale.LC_MONETARY, 'en_IN')
except locale.Error:
    try:
        locale.setlocale(locale.LC_MONETARY, 'en_US.UTF-8') # Fallback for some systems
    except locale.Error:
        pass # Fallback if locale is not available

# --- NEW: Helper function for PyInstaller ---
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Not bundled, use the script's directory
        base_path = os.path.abspath(os.path.dirname(sys.argv[0]))    
    return os.path.join(base_path, relative_path)


# --- THREAD 1: INITIALIZATION (Auto-Start) ---
def init_browser_thread():
    global automation_process, active_driver
    gui_queue.put(("status", "Checking System..."))

    if not os.path.exists(DRIVER_PATH):
        gui_queue.put(("error", f"Driver missing!\nPut 'msedgedriver.exe' here:\n{BASE_DIR}"))
        return

    service = Service(executable_path=DRIVER_PATH)
    edge_options = Options()
    edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

    # 1. CHECK FOR EXISTING BROWSER
    try:
        driver = webdriver.Edge(service=service, options=edge_options)
        gui_queue.put(("light", "green"))
        gui_queue.put(("status", "System Ready (Instant)"))
        return
    except:
        pass

        # 2. LAUNCH INVISIBLE BROWSER
    gui_queue.put(("status", "Starting Service..."))

    if os.path.exists(TEMP_PROFILE_DIR):
        try:
            shutil.rmtree(TEMP_PROFILE_DIR)
        except:
            pass

    cmd = [
        BROWSER_EXE, "--remote-debugging-port=9222", f"--user-data-dir={TEMP_PROFILE_DIR}",
        "--no-first-run", "--no-default-browser-check", "--ignore-certificate-errors",
        "--headless=new", "--disable-gpu", "--window-size=1920,1080",
        "--proxy-bypass-list=<-loopback>", TARGET_URL
    ]

    try:
        automation_process = subprocess.Popen(cmd, creationflags=0x08000000)
    except:
        automation_process = subprocess.Popen(cmd)

    # 3. CONNECTION LOOP
    connected = False
    for i in range(30):
        try:
            driver = webdriver.Edge(service=service, options=edge_options)
            connected = True
            break
        except:
            time.sleep(1)

    if not connected:
        gui_queue.put(("error", "Could not connect. Restart App."))
        return

    # 4. STABILIZATION
    gui_queue.put(("light", "orange"))
    for i in range(15, 0, -1):
        gui_queue.put(("status", f"Stabilizing... {i}s"))
        time.sleep(1)

    gui_queue.put(("light", "green"))
    gui_queue.put(("status", "System Ready"))


# --- THREAD 2: SEARCH ---
def search_thread(part_number):
    global active_driver
    try:
        service = Service(executable_path=DRIVER_PATH)
        edge_options = Options()
        edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

        gui_queue.put(("status", "Processing..."))

        try:
            driver = webdriver.Edge(service=service, options=edge_options)
            active_driver = driver
        except Exception as e:
            gui_queue.put(("error", "Connection Lost. Restart App."))
            gui_queue.put(("light", "red"))
            return

        if "heromotocorp" not in driver.current_url:
            driver.get(TARGET_URL)

        # Remove Popup
        try:
            driver.execute_script("""
                var banner = document.querySelector('.cc-window');
                if (banner) { banner.remove(); }
                var buttons = document.getElementsByTagName('button');
                for (var i = 0; i < buttons.length; i++) {
                    if (buttons[i].textContent.includes('I Understand')) { buttons[i].click(); }
                }
            """)
        except:
            pass

        # Search
        try:
            WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "ppartno")))
        except:
            gui_queue.put(("error", "Page not loaded."))
            return

        part_input = driver.find_element(By.ID, "ppartno")
        part_input.clear()
        part_input.send_keys(part_number)

        search_btn = driver.find_element(By.ID, "pbtn")
        driver.execute_script("arguments[0].click();", search_btn)

        # Scrape
        try:
            WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "datatableoldpart")))
        except:
            gui_queue.put(("not_found", "No results found."))
            return

        try:
            table = driver.find_element(By.ID, "datatableoldpart")
            rows = table.find_elements(By.TAG_NAME, "tr")

            if len(rows) < 2:
                gui_queue.put(("not_found", "Part not found."))
                return

            cols = rows[1].find_elements(By.TAG_NAME, "td")
            # Extract Data
            data = {
                "Part No": cols[1].text if len(cols) > 1 else "-",
                "Description": cols[3].text if len(cols) > 3 else "-",
                "MOQ": cols[5].text if len(cols) > 5 else "-",
                "Price": cols[6].text if len(cols) > 6 else "-"
            }
            gui_queue.put(("success", data))

        except Exception as e:
            gui_queue.put(("error", f"Read Error: {e}"))

    except Exception as e:
        gui_queue.put(("error", f"Error: {e}"))

class Database:
    """Handles all database operations for the order system."""
    def __init__(self, db_file):
        """Initializes the database connection and creates/updates tables."""
        self.db_file = db_file # <-- ADD THIS LINE
        self.conn = sqlite3.connect(db_file)
        self.cursor = self.conn.cursor()
        self._create_tables()

    def _create_tables(self):
        """Creates or alters tables to match the new schema."""
        # --- Standard Order Tables ---
        self.cursor.execute("CREATE TABLE IF NOT EXISTS parties (id INTEGER PRIMARY KEY, name TEXT NOT NULL)")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS items (id INTEGER PRIMARY KEY, name TEXT NOT NULL, part_number TEXT, price REAL)")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS orders (id INTEGER PRIMARY KEY, order_number TEXT , party_id INTEGER, order_date TEXT, status TEXT, total_amount REAL, last_saved_date TEXT, FOREIGN KEY (party_id) REFERENCES parties (id))")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS order_items (id INTEGER PRIMARY KEY, order_id INTEGER, item_id INTEGER, quantity INTEGER, unit_price REAL, vehicle TEXT, brand TEXT, moq TEXT, dlp REAL, FOREIGN KEY (order_id) REFERENCES orders (id), FOREIGN KEY (item_id) REFERENCES items (id))")

        # --- Accounting Tables ---
        self.cursor.execute("CREATE TABLE IF NOT EXISTS accounting_companies (id INTEGER PRIMARY KEY, name TEXT UNIQUE NOT NULL)")
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS invoices (id INTEGER PRIMARY KEY, company_id INTEGER, kar_id TEXT UNIQUE, customer_code TEXT, order_number TEXT, invoice_number TEXT UNIQUE, payment_mode TEXT, amount REAL, status TEXT, invoice_date TEXT, payment_date TEXT, partial_payment_date TEXT, debit_bank_name TEXT, account_number TEXT, transaction_ref TEXT, reference_id TEXT, FOREIGN KEY (company_id) REFERENCES accounting_companies (id))
        """)

        # --- Part Request Table ---
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS part_requests (id INTEGER PRIMARY KEY, request_id TEXT UNIQUE, customer_name TEXT, phone_number TEXT, security_amount REAL, payment_type TEXT, part_details TEXT, request_date TEXT, status TEXT)
        """)
        
        # --- NEW: Sales Commission Tables ---
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS sales_commissions (
                id INTEGER PRIMARY KEY,
                commission_no TEXT UNIQUE NOT NULL,
                mechanic_name TEXT,
                mobile_number TEXT,
                invoice_no TEXT,
                issue_date TEXT,
                status TEXT,
                total_amount REAL
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS commission_items (
                id INTEGER PRIMARY KEY,
                commission_id INTEGER,
                description TEXT,
                base_amount REAL,
                percentage REAL,
                rupees REAL,
                final_amount REAL,
                FOREIGN KEY (commission_id) REFERENCES sales_commissions (id)
            )
        """)

        # Add columns if they don't exist for backward compatibility
        try:
            self.cursor.execute("ALTER TABLE items ADD COLUMN part_number TEXT")
            self.cursor.execute("ALTER TABLE orders ADD COLUMN order_number TEXT UNIQUE");
            self.cursor.execute("ALTER TABLE orders ADD COLUMN status TEXT");
            self.cursor.execute("ALTER TABLE orders ADD COLUMN last_saved_date TEXT")
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN vehicle TEXT");
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN brand TEXT")
            # --- NEW MIGRATIONS ---
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN moq TEXT")
            self.cursor.execute("ALTER TABLE order_items ADD COLUMN dlp REAL")
        except sqlite3.OperationalError:
            pass
        self.conn.commit()
        
    def get_company_summary_by_month(self, month_str, year_str):
        """
        Calculates total, paid, and unpaid amounts per company for a specific month and year.
        Orders them by the highest total amount.
        """
        date_pattern = f'%-{month_str}-{year_str}'
        sql = """
        SELECT
            c.name,
            COUNT(i.id),
            SUM(CASE WHEN i.status = 'PAID' THEN i.amount ELSE 0 END),
            SUM(CASE WHEN i.status = 'UNPAID' THEN i.amount ELSE 0 END),
            SUM(i.amount) as total_amount
        FROM 
            invoices i
        JOIN 
            accounting_companies c ON i.company_id = c.id
        WHERE 
            i.invoice_date LIKE ?
        GROUP BY
            c.name
        ORDER BY
            total_amount DESC
        """
        try:
            self.cursor.execute(sql, (date_pattern,))
            return self.cursor.fetchall()
        except Exception as e:
            print(f"Error in get_company_summary_by_month: {e}")
            return []
        
    # --- Part Request Methods ---
    def generate_request_id(self):
        self.cursor.execute("SELECT request_id FROM part_requests ORDER BY id DESC LIMIT 1")
        last_id = self.cursor.fetchone()
        if last_id and last_id[0] and last_id[0].startswith("PAR"):
            try: num = int(last_id[0][3:]) + 1; return f"PAR{num}"
            except ValueError: return "PAR1"
        return "PAR1"

    def get_accounting_summary(self, month_str, year_str):
        """
        Calculates total, paid, and unpaid amounts for a specific month and year.
        Dates are stored as 'dd-mm-yyyy', so we use LIKE.
        """
        # Create a search pattern like '%-10-2025'
        date_pattern = f'%-{month_str}-{year_str}'
        sql = """
        SELECT
            SUM(amount),
            SUM(CASE WHEN status = 'PAID' THEN amount ELSE 0 END),
            SUM(CASE WHEN status = 'UNPAID' THEN amount ELSE 0 END)
        FROM 
            invoices
        WHERE 
            invoice_date LIKE ?
        """
        try:
            self.cursor.execute(sql, (date_pattern,))
            return self.cursor.fetchone()
        except Exception as e:
            print(f"Error in get_accounting_summary: {e}")
            return (0, 0, 0)

    
    def add_part_request(self, data):
        sql = "INSERT INTO part_requests (request_id, customer_name, phone_number, security_amount, payment_type, part_details, request_date, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
        values = (
            data['request_id'], data['customer_name'], data['phone_number'],
            data['security_amount'], data['payment_type'], data['part_details'],
            data['request_date'], data['status']
        )
        self.cursor.execute(sql, values)
        self.conn.commit()

    def update_part_request(self, request_id, data):
        sql = "UPDATE part_requests SET customer_name=?, phone_number=?, security_amount=?, payment_type=?, part_details=?, request_date=?, status=? WHERE id=?"
        values = (
            data['customer_name'], data['phone_number'], data['security_amount'],
            data['payment_type'], data['part_details'], data['request_date'],
            data['status'], request_id
        )
        self.cursor.execute(sql, values)
        self.conn.commit()

    def delete_part_request(self, request_id):
        self.cursor.execute("DELETE FROM part_requests WHERE id=?", (request_id,)); self.conn.commit()

    def get_all_part_requests(self):
        self.cursor.execute("SELECT * FROM part_requests ORDER BY id DESC"); return self.cursor.fetchall()

    def search_part_requests(self, order_id="", phone=""):
        if order_id:
            self.cursor.execute("SELECT * FROM part_requests WHERE request_id LIKE ? ORDER BY id DESC", (f'%{order_id}%',))
        elif phone:
            self.cursor.execute("SELECT * FROM part_requests WHERE phone_number LIKE ? ORDER BY id DESC", (f'%{phone}%',))
        return self.cursor.fetchall()

    # --- Accounting Methods ---
    def get_accounting_companies(self):
        self.cursor.execute("SELECT id, name FROM accounting_companies ORDER BY name"); return self.cursor.fetchall()

    def add_accounting_company(self, name):
        try: self.cursor.execute("INSERT INTO accounting_companies (name) VALUES (?)", (name,)); self.conn.commit(); return self.cursor.lastrowid
        except sqlite3.IntegrityError: return None

    def delete_accounting_company(self, company_id):
        self.cursor.execute("DELETE FROM invoices WHERE company_id=?", (company_id,)); self.cursor.execute("DELETE FROM accounting_companies WHERE id=?", (company_id,)); self.conn.commit()

    def get_all_invoices(self):
        sql = "SELECT i.id, ac.name, i.kar_id, i.customer_code, i.order_number, i.invoice_number, i.payment_mode, i.amount, i.status, i.invoice_date, i.payment_date, i.partial_payment_date, i.debit_bank_name, i.account_number, i.transaction_ref, i.reference_id FROM invoices i JOIN accounting_companies ac ON i.company_id = ac.id ORDER BY i.id DESC"
        self.cursor.execute(sql); return self.cursor.fetchall()

    def add_invoice(self, data):
        sql = "INSERT INTO invoices (company_id, kar_id, customer_code, order_number, invoice_number, payment_mode, amount, status, invoice_date, payment_date, partial_payment_date, debit_bank_name, account_number, transaction_ref, reference_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        values = (
            data['company_id'], data['kar_id'], data['customer_code'], data['order_number'],
            data['invoice_number'], data['payment_mode'], data['amount'], data['status'],
            data['invoice_date'], data['payment_date'], data['partial_payment_date'],
            data['debit_bank_name'], data['account_number'], data['transaction_ref'],
            data['reference_id']
        )
        self.cursor.execute(sql, values)
        self.conn.commit()

    def search_invoices_by_number(self, invoice_number):
        """Searches for invoices matching a given invoice number."""
        sql = "SELECT i.id, ac.name, i.kar_id, i.customer_code, i.order_number, i.invoice_number, i.payment_mode, i.amount, i.status, i.invoice_date, i.payment_date, i.partial_payment_date, i.debit_bank_name, i.account_number, i.transaction_ref, i.reference_id FROM invoices i JOIN accounting_companies ac ON i.company_id = ac.id WHERE i.invoice_number LIKE ? ORDER BY i.id DESC"
        self.cursor.execute(sql, (f'%{invoice_number}%',))
        return self.cursor.fetchall()

    def search_invoices_by_ref(self, search_term):
        """Searches for invoices by transaction_ref or reference_id."""
        try:
            # Use a wildcard search to find partial matches
            wildcard_term = f'%{search_term}%'
            
            # This query selects the same columns as get_all_invoices
            sql = """
            SELECT 
                i.id, ac.name, i.kar_id, i.customer_code, i.order_number, i.invoice_number, 
                i.payment_mode, i.amount, i.status, i.invoice_date, i.payment_date, 
                i.partial_payment_date, i.debit_bank_name, i.account_number, 
                i.transaction_ref, i.reference_id 
            FROM 
                invoices i 
            JOIN 
                accounting_companies ac ON i.company_id = ac.id
            WHERE 
                i.transaction_ref LIKE ? OR i.reference_id LIKE ?
            ORDER BY 
                i.id DESC
            """
            self.cursor.execute(sql, (wildcard_term, wildcard_term))
            return self.cursor.fetchall()
        except sqlite3.Error as e: 
            print(f"Database error in search_invoices_by_ref: {e}")
            return []
    # --- END OF NEW METHOD ---

    def delete_invoice(self, invoice_id):
        self.cursor.execute("DELETE FROM invoices WHERE id=?", (invoice_id,)); self.conn.commit()

    def check_invoice_number_exists(self, invoice_number, current_id=None):
        """Checks if an invoice number exists, optionally ignoring the current record."""
        sql = "SELECT id FROM invoices WHERE invoice_number=?"
        params = [invoice_number]

        # This 'if' block is the new part that uses current_id
        if current_id is not None:
            sql += " AND id != ?"
            params.append(current_id)

        self.cursor.execute(sql, params)
        return self.cursor.fetchone() is not None

    def get_invoice_details_by_id(self, invoice_id):
        """Gets all data for a single invoice, joining to get the company name."""
        # Join with companies to get the name, return all other fields
        sql = """
              SELECT ac.name as company_name, \
                     i.kar_id, \
                     i.customer_code, \
                     i.order_number, \
                     i.invoice_number, \
                     i.payment_mode, \
                     i.amount, \
                     i.status, \
                     i.invoice_date, \
                     i.payment_date, \
                     i.partial_payment_date, \
                     i.debit_bank_name, \
                     i.account_number, \
                     i.transaction_ref, \
                     i.reference_id, \
                     i.company_id
              FROM invoices i \
                       JOIN \
                   accounting_companies ac ON i.company_id = ac.id
              WHERE i.id = ? \
              """
        self.cursor.execute(sql, (invoice_id,))
        row = self.cursor.fetchone()
        if not row:
            return None

        # Get column names
        cols = [description[0] for description in self.cursor.description]
        # Return as a dictionary
        return dict(zip(cols, row))

    def update_invoice(self, invoice_id, data):
        """Updates an existing invoice in the database."""
        sql = """
              UPDATE invoices
              SET company_id           = ?, \
                  kar_id               = ?, \
                  customer_code        = ?, \
                  order_number         = ?,
                  invoice_number       = ?, \
                  payment_mode         = ?, \
                  amount               = ?, \
                  status               = ?,
                  invoice_date         = ?, \
                  payment_date         = ?, \
                  partial_payment_date = ?,
                  debit_bank_name      = ?, \
                  account_number       = ?, \
                  transaction_ref      = ?,
                  reference_id         = ?
              WHERE id = ? \
              """
        values = (
            data['company_id'], data['kar_id'], data['customer_code'], data['order_number'],
            data['invoice_number'], data['payment_mode'], data['amount'], data['status'],
            data['invoice_date'], data['payment_date'], data['partial_payment_date'],
            data['debit_bank_name'], data['account_number'], data['transaction_ref'],
            data['reference_id'],
            invoice_id  # The last parameter is the ID for the WHERE clause
        )
        self.cursor.execute(sql, values)
        self.conn.commit()


    def generate_kar_id(self):
        self.cursor.execute("SELECT kar_id FROM invoices ORDER BY id DESC LIMIT 1")
        last_id = self.cursor.fetchone()
        if last_id and last_id[0] and last_id[0].startswith("KAR"):
            try: num = int(last_id[0][3:]) + 1; return f"KAR{num}"
            except ValueError: return "KAR1"
        return "KAR1"

    # --- Standard Order Methods ---
    def generate_order_number(self):
        now = datetime.now(); month_codes = {1: 'JA', 2: 'FE', 3: 'MR', 4: 'AP', 5: 'MY', 6: 'JN', 7: 'JL', 8: 'AU', 9: 'SE', 10: 'OC', 11: 'NV', 12: 'DE'}
        month_code = month_codes[now.month]; prefix = f"OR{month_code}"
        self.cursor.execute("SELECT order_number FROM orders WHERE order_number LIKE ? ORDER BY order_number DESC LIMIT 1", (f'{prefix}%',))
        last_order = self.cursor.fetchone()
        if last_order and last_order[0]:
            try: last_num = int(last_order[0][-2:]); new_num = last_num + 1
            except (ValueError, TypeError): new_num = 1
        else: new_num = 1
        return f"{prefix}{new_num:02d}"

    def get_suggestions(self, table, query):
        self.cursor.execute(f"SELECT name FROM {table} WHERE name LIKE ?", (f'%{query}%',)); return [row[0] for row in self.cursor.fetchall()]

    def get_or_create_id(self, table, name, details=None):
        # If we are saving an item, check BOTH name AND part number
        if table == 'items' and details:
            part_no = details.get('part_number', '')
            self.cursor.execute("SELECT id FROM items WHERE name=? AND part_number=?", (name, part_no))
        else:
            # For parties, just matching name is fine
            self.cursor.execute(f"SELECT id FROM {table} WHERE name=?", (name,))

        result = self.cursor.fetchone()

        if result:
            # Returns ID only if BOTH name and part number match exactly
            return result[0]
        else:
            # Create a new entry if that specific combination doesn't exist
            if details is None: details = {}
            if table == 'parties':
                self.cursor.execute("INSERT INTO parties (name) VALUES (?)", (name,))
            elif table == 'items':
                self.cursor.execute("INSERT INTO items (name, part_number, price) VALUES (?, ?, ?)",
                                    (name, details.get('part_number', ''), details.get('price', 0.0)))
            self.conn.commit()
            return self.cursor.lastrowid

    def get_party_id_by_name(self, name):
        self.cursor.execute("SELECT id FROM parties WHERE name=?", (name,)); result = self.cursor.fetchone(); return result[0] if result else None

    def get_all_orders_for_party(self, party_id):
        self.cursor.execute("SELECT id, order_number, status FROM orders WHERE party_id = ? ORDER BY order_date DESC", (party_id,)); return self.cursor.fetchall()

    def get_order_id_by_number(self, order_number):
        self.cursor.execute("SELECT id FROM orders WHERE order_number=?", (order_number,)); result = self.cursor.fetchone(); return result[0] if result else None

    def get_current_order_info_for_party(self, party_id):
        self.cursor.execute("SELECT id, order_number FROM orders WHERE party_id = ? AND status = 'Current' LIMIT 1", (party_id,)); return self.cursor.fetchone()

    def get_order_items(self, order_id):
        sql = "SELECT i.name, i.part_number, oi.quantity, oi.unit_price, oi.vehicle, oi.brand, oi.moq, oi.dlp FROM order_items oi JOIN items i ON oi.item_id = i.id WHERE oi.order_id = ?"
        self.cursor.execute(sql, (order_id,));
        return self.cursor.fetchall()

    def save_or_update_order(self, order_id, party_id, total_amount, items_data, status, pregen_order_number=None):
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S"); update_items = (status == 'Current')
        if order_id:
            self.cursor.execute("UPDATE orders SET last_saved_date=?, total_amount=?, status=? WHERE id=?", (now_str, total_amount, status, order_id))
            if update_items:
                self.cursor.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
                # Modified Insert
                sql_items = "INSERT INTO order_items (order_id, item_id, quantity, unit_price, vehicle, brand, moq, dlp) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
                for item in items_data:
                    self.cursor.execute(sql_items, (order_id, item['item_id'], item['quantity'], item['unit_price'], item.get('vehicle', ''), item.get('brand', ''), item.get('moq', '-'), item.get('dlp', 0.0)))
        else:
            order_number = pregen_order_number or self.generate_order_number()
            self.cursor.execute("INSERT INTO orders (order_number, party_id, order_date, last_saved_date, total_amount, status) VALUES (?, ?, ?, ?, ?, ?)", (order_number, party_id, now_str, now_str, total_amount, status))
            order_id = self.cursor.lastrowid
            # Modified Insert
            sql_items = "INSERT INTO order_items (order_id, item_id, quantity, unit_price, vehicle, brand, moq, dlp) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
            for item in items_data:
                 self.cursor.execute(sql_items, (order_id, item['item_id'], item['quantity'], item['unit_price'], item.get('vehicle', ''), item.get('brand', ''), item.get('moq', '-'), item.get('dlp', 0.0)))
        self.conn.commit(); return order_id

    def get_all_orders(self):
        sql = "SELECT o.order_number, p.name, o.order_date, o.last_saved_date, o.status FROM orders o JOIN parties p ON o.party_id = p.id ORDER BY o.last_saved_date DESC"
        self.cursor.execute(sql); return self.cursor.fetchall()

    def search_orders_by_part_number(self, part_number):
        """Finds all orders that contain a specific part number."""
        sql = """
            SELECT DISTINCT o.order_number, p.name, o.order_date, o.last_saved_date, o.status
            FROM orders o
            JOIN parties p ON o.party_id = p.id
            JOIN order_items oi ON o.id = oi.order_id
            JOIN items i ON oi.item_id = i.id
            WHERE i.part_number LIKE ?
            ORDER BY o.last_saved_date DESC
        """
        self.cursor.execute(sql, (f'%{part_number}%',))
        return self.cursor.fetchall()

    def delete_order(self, order_id):
        """Deletes an order and all its associated items."""
        self.cursor.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        self.cursor.execute("DELETE FROM orders WHERE id=?", (order_id,))
        self.conn.commit()

    # --- NEW: Sales Commission Methods ---
    def generate_commission_no(self):
        self.cursor.execute("SELECT commission_no FROM sales_commissions ORDER BY id DESC LIMIT 1")
        last_id = self.cursor.fetchone()
        if last_id and last_id[0] and last_id[0].startswith("FAZ"):
            try: num = int(last_id[0][3:]) + 1; return f"FAZ{num:02d}"
            except ValueError: return "FAZ01"
        return "FAZ01"

    def save_commission(self, main_data, items_data):
        sql_main = "INSERT INTO sales_commissions (commission_no, mechanic_name, mobile_number, invoice_no, issue_date, status, total_amount) VALUES (?, ?, ?, ?, ?, ?, ?)"
        self.cursor.execute(sql_main, (
            main_data['commission_no'], main_data['mechanic_name'], main_data['mobile_number'],
            main_data['invoice_no'], main_data['issue_date'], main_data['status'], main_data['total_amount']
        ))
        commission_id = self.cursor.lastrowid
        
        sql_item = "INSERT INTO commission_items (commission_id, description, base_amount, percentage, rupees, final_amount) VALUES (?, ?, ?, ?, ?, ?)"
        for item in items_data:
            self.cursor.execute(sql_item, (
                commission_id, item['description'], item['base_amount'],
                item['percentage'], item['rupees'], item['final_amount']
            ))
        self.conn.commit()
        return commission_id

    def get_all_commissions(self):
        self.cursor.execute("SELECT id, commission_no, mechanic_name, invoice_no, issue_date, status, total_amount FROM sales_commissions ORDER BY id DESC")
        return self.cursor.fetchall()

    def get_commission_details(self, commission_id):
        self.cursor.execute("SELECT * FROM sales_commissions WHERE id=?", (commission_id,))
        main_details = self.cursor.fetchone()
        if not main_details:
            return None, []
            
        # Remap tuple to dict
        cols = ["id", "commission_no", "mechanic_name", "mobile_number", "invoice_no", "issue_date", "status", "total_amount"]
        main_dict = dict(zip(cols, main_details))
        
        self.cursor.execute("SELECT description, base_amount, percentage, rupees, final_amount FROM commission_items WHERE commission_id=?", (commission_id,))
        items = self.cursor.fetchall()
        
        item_cols = ["description", "base_amount", "percentage", "rupees", "final_amount"]
        items_list = [dict(zip(item_cols, item)) for item in items]
        
        return main_dict, items_list

    def update_commission(self, commission_id, main_data, items_data):
        """Updates an existing commission slip and its items."""
        sql_main = """
            UPDATE sales_commissions 
            SET mechanic_name=?, mobile_number=?, invoice_no=?, issue_date=?, status=?, total_amount=?
            WHERE id=?
        """
        self.cursor.execute(sql_main, (
            main_data['mechanic_name'], main_data['mobile_number'],
            main_data['invoice_no'], main_data['issue_date'], main_data['status'], 
            main_data['total_amount'], commission_id
        ))
        
        # Re-delete and add items
        self.cursor.execute("DELETE FROM commission_items WHERE commission_id=?", (commission_id,))
        
        sql_item = "INSERT INTO commission_items (commission_id, description, base_amount, percentage, rupees, final_amount) VALUES (?, ?, ?, ?, ?, ?)"
        for item in items_data:
            self.cursor.execute(sql_item, (
                commission_id, item['description'], item['base_amount'],
                item['percentage'], item['rupees'], item['final_amount']
            ))
        self.conn.commit()
        return commission_id

    def delete_commission(self, commission_id):
        self.cursor.execute("DELETE FROM commission_items WHERE commission_id=?", (commission_id,))
        self.cursor.execute("DELETE FROM sales_commissions WHERE id=?", (commission_id,))
        self.conn.commit()

    def search_commissions(self, commission_no="", mobile_number=""):
        """Searches commissions by commission_no or mobile_number."""
        sql = "SELECT id, commission_no, mechanic_name, invoice_no, issue_date, status, total_amount FROM sales_commissions"
        params = []
        if commission_no:
            sql += " WHERE commission_no LIKE ?"
            params.append(f'%{commission_no}%')
        elif mobile_number:
            sql += " WHERE mobile_number LIKE ?"
            params.append(f'%{mobile_number}%')
        
        sql += " ORDER BY id DESC"
        self.cursor.execute(sql, params)
        return self.cursor.fetchall()

    def __del__(self):
        if self.conn: self.conn.close()

# --- Autocomplete Widget ---
class AutocompleteEntry(ttk.Entry):
    def __init__(self, master, table, db, **kwargs):
        if "textvariable" in kwargs: self.var = kwargs["textvariable"]
        else: self.var = tk.StringVar(); kwargs["textvariable"] = self.var
        super().__init__(master, **kwargs)
        self.table = table; self.db = db; self.trace_active = True; self.listbox = None
        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection); self.bind("<Return>", self.selection)
        self.bind("<Up>", self.move_up); self.bind("<Down>", self.move_down)

    def changed(self, name, index, mode):
        if not self.trace_active: return
        if self.var.get() == '':
            if self.listbox: self.listbox.destroy(); self.listbox = None
        else:
            words = self.db.get_suggestions(self.table, self.var.get())
            if words:
                if not self.listbox:
                    self.listbox = tk.Listbox(self.master, width=self["width"], height=4); self.listbox.place(x=self.winfo_x(), y=self.winfo_y() + self.winfo_height())
                    self.listbox.bind("<Double-Button-1>", self.selection); self.listbox.bind("<Right>", self.selection)
                self.listbox.delete(0, tk.END)
                for w in words: self.listbox.insert(tk.END, w)
            else:
                if self.listbox: self.listbox.destroy(); self.listbox = None

    def selection(self, event=None):
        if self.listbox:
            try:
                self.var.set(self.listbox.get(self.listbox.curselection()))
                self.listbox.destroy()
                self.listbox = None
                self.icursor(tk.END)
            except tk.TclError:
                pass # Ignore if no selection

    def set_text_programmatically(self, text):
        self.trace_active = False; self.var.set(text); self.trace_active = True
        if self.listbox: self.listbox.destroy(); self.listbox = None

    def move_up(self, event):
        if self.listbox:
            if not self.listbox.curselection(): self.listbox.selection_set(0)
            index = int(self.listbox.curselection()[0])
            if index > 0: self.listbox.selection_clear(index); index -= 1; self.listbox.selection_set(index); self.listbox.activate(index)

    def move_down(self, event):
        if self.listbox:
            if not self.listbox.curselection(): self.listbox.selection_set(0)
            index = int(self.listbox.curselection()[0])
            if index < self.listbox.size() - 1: self.listbox.selection_clear(index); index += 1; self.listbox.selection_set(index); self.listbox.activate(index)

# --- Utility Function ---
def format_currency(amount):
    """Safely formats a number to a string with two decimal places."""
    try:
        # Format as a float with 2 decimal places. No symbols, no commas.
        return f"{float(amount):.2f}"
    except (ValueError, TypeError):
        # Return a simple default
        return "0.00"

def safe_float(value, default=0.0):
    """Safely converts a value to float."""
    try:
        # Handle potential locale issues (e.g., currency symbols, commas)
        if isinstance(value, str):
            # Remove currency symbol and grouping separators
            cleaned_value = value.strip()
            if locale.localeconv()['currency_symbol']:
                 cleaned_value = cleaned_value.replace(locale.localeconv()['currency_symbol'], '')
            if locale.localeconv()['grouping']:
                 cleaned_value = cleaned_value.replace(locale.localeconv()['thousands_sep'], '')
            # Use locale.atof for locale-aware conversion
            return locale.atof(cleaned_value)
        else:
            return float(value) # Try direct float conversion
    except (ValueError, TypeError, AttributeError): # Added AttributeError
        return default


# --- Application Pages ---
class OrderPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent);
        self.db = db;
        self.app = app
        self.items_in_order = []
        self.current_order_id = None;
        self.current_party_id = None;
        self.pending_order_number = None
        self._setup_widgets()

    def _setup_widgets(self):
        page_controls = ttk.Frame(self, padding=(10, 10, 10, 0));
        page_controls.pack(fill=tk.X)
        ttk.Button(page_controls, text="Start New Order", command=self.start_new_order).pack(side=tk.LEFT)

        self.top_frame = ttk.Frame(self, padding=10);
        self.top_frame.pack(fill=tk.X)

        # Simple Autocomplete Party
        party_frame = ttk.LabelFrame(self.top_frame, text="Party", padding=5);
        party_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.party_var = tk.StringVar()
        self.party_entry = AutocompleteEntry(party_frame, 'parties', self.db, textvariable=self.party_var, width=40)
        self.party_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(party_frame, text="Save", command=self.save_party).pack(side=tk.RIGHT, padx=(5, 0))
        self.party_entry.bind("<Return>", self.load_party_data)
        self.party_entry.bind("<FocusOut>", self.load_party_data)

        order_frame = ttk.LabelFrame(self.top_frame, text="Select Order", padding=5);
        order_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        self.order_var = tk.StringVar()
        self.order_combo = ttk.Combobox(order_frame, textvariable=self.order_var, state='readonly', width=30)
        self.order_combo.pack(fill=tk.X, expand=True);
        self.order_combo.bind('<<ComboboxSelected>>', self.on_order_selected)

        self._create_item_widgets();
        self._create_bill_details_widgets()

    def _create_item_widgets(self):
        self.item_frame = ttk.LabelFrame(self, text="Add Item", padding=10);
        self.item_frame.pack(fill=tk.X, padx=10, pady=5)
        for i in range(12): self.item_frame.columnconfigure(i, weight=1)

        ttk.Label(self.item_frame, text="Part No:").grid(row=0, column=0, sticky=tk.W)
        self.item_part_no_var = tk.StringVar()
        self.item_part_no_var.trace("w", self._force_uppercase)
        self.part_no_entry = ttk.Entry(self.item_frame, textvariable=self.item_part_no_var, width=15)
        self.part_no_entry.grid(row=0, column=1, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Qty:").grid(row=0, column=2, sticky=tk.W)
        self.item_qty_var = tk.IntVar(value=1)
        self.qty_entry = ttk.Entry(self.item_frame, textvariable=self.item_qty_var, width=5)
        self.qty_entry.grid(row=0, column=3, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Item Name:").grid(row=0, column=4, sticky=tk.W)
        self.item_name_var = tk.StringVar()
        self.item_name_entry = ttk.Entry(self.item_frame, textvariable=self.item_name_var, width=20)
        self.item_name_entry.grid(row=0, column=5, sticky=tk.EW, padx=2)

        self.add_item_button = ttk.Button(self.item_frame, text="Add", command=self.add_item_to_order)
        self.add_item_button.grid(row=0, column=6, padx=5)

        self.part_no_entry.bind("<Return>", lambda e: self.qty_entry.focus())
        self.qty_entry.bind("<Return>", lambda e: self.item_name_entry.focus())
        self.item_name_entry.bind("<Return>", lambda e: self.add_item_to_order())

    def _create_bill_details_widgets(self):
        self.bill_list_frame = ttk.LabelFrame(self, text="Current Bill Details", padding=10);
        self.bill_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ("sr_no", "part_no", "item_name", "qty")
        headers = {"sr_no": "SN", "part_no": "Part No", "qty": "Qty", "item_name": "Item Name"}
        widths = {"sr_no": 40, "part_no": 120, "qty": 60, "item_name": 300}

        self.tree = ttk.Treeview(self.bill_list_frame, columns=cols, show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)
        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor=tk.CENTER if col != "item_name" else tk.W)

        total_frame = ttk.Frame(self, padding=10);
        total_frame.pack(fill=tk.X)
        status_frame = ttk.Frame(total_frame);
        status_frame.pack(side=tk.LEFT)
        ttk.Label(status_frame, text="Order Status:").pack(side=tk.LEFT)
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(status_frame, textvariable=self.status_var, values=['Current', 'Sended'],
                                         state='readonly')
        self.status_combo.pack(side=tk.LEFT, padx=5);
        self.status_combo.bind('<<ComboboxSelected>>', self.on_status_change)

        self.delete_item_button = ttk.Button(total_frame, text="Delete Selected", command=self.delete_selected_item)
        self.delete_item_button.pack(side=tk.RIGHT, padx=2)
        self.save_button = ttk.Button(total_frame, text="Save Order", command=self.save_order)
        self.save_button.pack(side=tk.RIGHT, padx=2)

    def _force_uppercase(self, *args):
        val = self.item_part_no_var.get()
        if val != val.upper(): self.item_part_no_var.set(val.upper())

    def add_item_to_order(self):
        name = self.item_name_var.get();
        part_no = self.item_part_no_var.get().strip()
        if self._check_duplicate_part_no(part_no): return
        try:
            qty = self.item_qty_var.get()
        except:
            qty = 0
        if not name or qty <= 0: messagebox.showwarning("Input Error", "Enter Name and Quantity."); return

        self.items_in_order.append({
            'name': name, 'part_no': part_no, 'qty': qty,
            'price': 0.0, 'vehicle': '', 'brand': '', 'moq': '-', 'dlp': 0.0
        })
        self.refresh_bill_treeview();
        self.clear_item_fields();
        self.part_no_entry.focus()

    def refresh_bill_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for i, item in enumerate(self.items_in_order):
            self.tree.insert("", tk.END, values=(i + 1, item['part_no'], item['name'], item['qty']))

    def start_new_order(self):
        if hasattr(self, 'party_entry'): self.party_entry.set_text_programmatically('')
        self.order_var.set('');
        self.order_combo['values'] = [];
        self.current_party_id = None;
        self.clear_bill_details();
        if hasattr(self, 'party_entry'):
            self.party_entry.focus()
        else:
            self.part_no_entry.focus()

    def save_party(self):
        party_name = self.party_var.get().strip()
        if not party_name: return
        self.current_party_id = self.db.get_or_create_id('parties', party_name)
        messagebox.showinfo("Success", f"Party '{party_name}' saved.");
        self.load_party_data(force_reload=True)

    def load_party_data(self, event=None, force_reload=False):
        party_name = self.party_var.get().strip() if hasattr(self, 'party_var') else "Non OEM"
        if not party_name:
            if self.current_party_id is not None: self.current_party_id = None; self.clear_bill_details();
            self.order_combo['values'] = [];
            self.order_var.set('')
            return
        party_id = self.db.get_party_id_by_name(party_name)
        if party_id is None:
            self.current_party_id = None;
            return
        if party_id == self.current_party_id and not force_reload: return
        self.current_party_id = party_id
        all_orders = self.db.get_all_orders_for_party(self.current_party_id)
        display_list = [f"{order[1]} ({order[2]})" for order in all_orders]
        current_order_info = next((order for order in all_orders if order[2] == 'Current'), None)
        if current_order_info:
            order_id, order_number, _ = current_order_info
            self.load_order_details(order_id, f"{order_number} (Current)");
            self.order_var.set(f"{order_number} (Current)")
        else:
            self.clear_bill_details();
            self.pending_order_number = self.db.generate_order_number()
            self.item_frame.config(text=f"Items for Order: {self.pending_order_number}")
            self.status_var.set("Current");
            self.on_status_change();
            self.order_var.set("New Order (Current)")
        self.order_combo['values'] = display_list

    def on_order_selected(self, event=None):
        selected = self.order_var.get()
        if not selected: return
        if selected == "New Order (Current)":
            self.clear_bill_details();
            self.pending_order_number = self.db.generate_order_number()
            self.item_frame.config(text=f"Items for Order: {self.pending_order_number}")
            self.status_var.set("Current");
            self.on_status_change();
            return
        self.pending_order_number = None;
        order_number = selected.split(' ')[0]
        order_id = self.db.get_order_id_by_number(order_number)
        if order_id: self.load_order_details(order_id, selected)

    def load_order_details(self, order_id, order_str):
        self.clear_bill_details();
        self.current_order_id = order_id
        try:
            order_number, status_str = order_str.split(' '); status = status_str.strip('()')
        except ValueError:
            order_number = order_str; status = 'Current'
        self.item_frame.config(text=f"Items for Order: {order_number}");
        self.status_var.set(status)
        items = self.db.get_order_items(order_id)
        for name, part_no, qty, price, vehicle, brand, moq, dlp in items:
            self.items_in_order.append({
                'name': name, 'part_no': part_no, 'qty': qty,
                'price': price, 'vehicle': vehicle, 'brand': brand, 'moq': moq, 'dlp': dlp or 0.0
            })
        self.refresh_bill_treeview();
        self.on_status_change()

    def _check_duplicate_part_no(self, part_no):
        if part_no and part_no.strip():
            for item in self.items_in_order:
                if item.get('part_no') == part_no:
                    messagebox.showwarning("Duplicate", "Part already in order.");
                    return True
        return False

    def save_order(self):
        party_name = self.party_var.get() if hasattr(self, 'party_var') else "Non OEM"
        if not party_name: messagebox.showerror("Error", "Party Name required."); return
        self.current_party_id = self.db.get_or_create_id('parties', party_name)
        status = self.status_var.get()
        if not status: messagebox.showerror("Error", "Set order status."); return
        if status == 'Current' and self.current_order_id is None:
            if self.db.get_current_order_info_for_party(self.current_party_id):
                messagebox.showwarning("Blocked", "Party already has a Current order.");
                return
        items_to_save, total = [], 0
        for item in self.items_in_order:
            item_id = self.db.get_or_create_id('items', item['name'],
                                               {'part_number': item['part_no'], 'price': item['price']})
            items_to_save.append({
                'item_id': item_id, 'quantity': item['qty'], 'unit_price': item['price'],
                'vehicle': item.get('vehicle', ''), 'brand': item.get('brand', ''),
                'moq': item.get('moq', ''), 'dlp': item.get('dlp', 0.0)
            })
            total += item['qty'] * item['price']
        try:
            self.db.save_or_update_order(self.current_order_id, self.current_party_id, total, items_to_save, status,
                                         self.pending_order_number)
            messagebox.showinfo("Success", "Order saved!");
            self.load_party_data(force_reload=True);
            self.app.history_page.refresh_orders()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_bill_details(self):
        self.clear_item_fields();
        self.items_in_order.clear();
        self.current_order_id = None;
        self.pending_order_number = None
        for i in self.tree.get_children(): self.tree.delete(i)
        self.status_var.set('');
        self.item_frame.config(text="Add Item to Order")

    def clear_item_fields(self):
        self.item_name_var.set('');
        self.item_part_no_var.set('');
        self.item_qty_var.set(1);
        self.part_no_entry.focus()

    def delete_selected_item(self):
        if not self.tree.selection(): return
        try:
            del self.items_in_order[int(self.tree.item(self.tree.selection()[0], 'values')[0]) - 1]
            self.refresh_bill_treeview()
        except:
            pass

    def on_status_change(self, event=None):
        state = 'normal' if self.status_var.get() == 'Current' else 'disabled'
        for child in self.item_frame.winfo_children():
            try:
                child.config(state=state)
            except:
                pass
        self.add_item_button.config(state=state);
        self.delete_item_button.config(state=state);
        self.save_button.config(state='normal')
### 2. The New `HeroOrderPage` (Fully Independent)

class HeroOrderPage(OrderPage):
    def __init__(self, parent, db, app):
        super().__init__(parent, db, app)

    def _setup_widgets(self):
        # --- HERO HEADER ---
        header = ttk.Frame(self, padding=10)
        header.pack(fill=tk.X)

        # Scraper Status in Top Right
        self.status_frame = ttk.Frame(header)
        self.status_frame.pack(side=tk.RIGHT)
        self.lbl_scraper_status = ttk.Label(self.status_frame, text="Starting Scraper...", font=("Arial", 9))
        self.lbl_scraper_status.pack(side=tk.LEFT, padx=5)
        self.canvas_light = tk.Canvas(self.status_frame, width=20, height=20, highlightthickness=0)
        self.canvas_light.pack(side=tk.LEFT, padx=5)
        self.light_circle = self.canvas_light.create_oval(2, 2, 18, 18, fill="red", outline="gray")

        ttk.Button(header, text="Start New Order", command=self.start_new_order).pack(side=tk.LEFT)
        ttk.Label(header, text="  [Hero Genuine Mode]", font=("Arial", 10, "bold"), foreground="blue").pack(
            side=tk.LEFT)

        # --- SELECTION FORM ---
        sel_frame = ttk.Frame(self, padding=10)
        sel_frame.pack(fill=tk.X)

        # Party (Disabled / Fixed)
        pf = ttk.LabelFrame(sel_frame, text="Party", padding=5)
        pf.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.party_var = tk.StringVar(value="Hero Genuine")
        ttk.Entry(pf, textvariable=self.party_var, state='disabled').pack(fill=tk.X)

        # Order Selector
        of = ttk.LabelFrame(sel_frame, text="Select Hero Order", padding=5)
        of.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        self.order_var = tk.StringVar()
        self.order_combo = ttk.Combobox(of, textvariable=self.order_var, state='readonly')
        self.order_combo.pack(fill=tk.X, expand=True)
        self.order_combo.bind('<<ComboboxSelected>>', self.on_order_selected)

        # Create Items and List
        self._create_item_widgets()
        self._create_bill_details_widgets()

        # Force Party Load
        self.current_party_id = self.db.get_or_create_id('parties', 'Hero Genuine')
        self.load_party_data(force_reload=True)

        # Scraper check loop
        self.process_queue()

    def _create_item_widgets(self):
        self.item_frame = ttk.LabelFrame(self, text="Add Hero Item", padding=10)
        self.item_frame.pack(fill=tk.X, padx=10, pady=5)

        # PART NO
        ttk.Label(self.item_frame, text="Part No:").grid(row=0, column=0)
        self.item_part_no_var = tk.StringVar()
        self.item_part_no_var.trace("w", self._force_uppercase)
        self.part_no_entry = ttk.Entry(self.item_frame, textvariable=self.item_part_no_var, width=15)
        self.part_no_entry.grid(row=0, column=1, padx=2)
        self.part_no_entry.bind("<FocusOut>", self.trigger_scraper)

        # QTY
        ttk.Label(self.item_frame, text="Qty:").grid(row=0, column=2)
        self.item_qty_var = tk.IntVar(value=1)
        self.qty_entry = ttk.Entry(self.item_frame, textvariable=self.item_qty_var, width=5)
        self.qty_entry.grid(row=0, column=3, padx=2)

        # ITEM NAME
        ttk.Label(self.item_frame, text="Item Name:").grid(row=0, column=4)
        self.item_name_var = tk.StringVar()
        self.item_name_entry = ttk.Entry(self.item_frame, textvariable=self.item_name_var, width=20)
        self.item_name_entry.grid(row=0, column=5, padx=2)

        # PRICE
        ttk.Label(self.item_frame, text="Price:").grid(row=0, column=6)
        self.item_price_var = tk.DoubleVar(value=0.0)
        self.price_entry = ttk.Entry(self.item_frame, textvariable=self.item_price_var, width=8)
        self.price_entry.grid(row=0, column=7, padx=2)

        # MOQ
        ttk.Label(self.item_frame, text="MOQ:").grid(row=0, column=8)
        self.item_moq_var = tk.StringVar()
        self.moq_entry = ttk.Entry(self.item_frame, textvariable=self.item_moq_var, width=6)
        self.moq_entry.grid(row=0, column=9, padx=2)

        # DLP
        ttk.Label(self.item_frame, text="DLP %:").grid(row=0, column=10)
        self.item_dlp_var = tk.DoubleVar(value=24.0)
        self.dlp_entry = ttk.Entry(self.item_frame, textvariable=self.item_dlp_var, width=5)
        self.dlp_entry.grid(row=0, column=11, padx=2)

        self.add_item_button = ttk.Button(self.item_frame, text="Add", command=self.add_item_to_order)
        self.add_item_button.grid(row=0, column=12, padx=5)

        # Navigation
        self.part_no_entry.bind("<Return>", lambda e: self.qty_entry.focus())
        self.qty_entry.bind("<Return>", lambda e: self.item_name_entry.focus())
        self.item_name_entry.bind("<Return>", lambda e: self.price_entry.focus())
        self.price_entry.bind("<Return>", lambda e: self.dlp_entry.focus())
        self.dlp_entry.bind("<Return>", lambda e: self.add_item_to_order())

    def _create_bill_details_widgets(self):
        self.bill_list_frame = ttk.LabelFrame(self, text="Current Bill Details", padding=10)
        self.bill_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ("sr_no", "part_no", "qty", "item_name", "price", "moq", "dlp", "total")
        self.tree = ttk.Treeview(self.bill_list_frame, columns=cols, show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)

        headers = {
            "sr_no": "SN", "part_no": "Part No", "qty": "Qty", "item_name": "Item Name",
            "price": "Price", "moq": "MOQ", "dlp": "DLP%", "total": "Exp. Total"
        }
        widths = {
            "sr_no": 40, "part_no": 100, "qty": 40, "item_name": 200,
            "price": 80, "moq": 60, "dlp": 60, "total": 100
        }
        for col in cols:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor=tk.CENTER if col != "item_name" else tk.W)

        # Bottom
        total_frame = ttk.Frame(self, padding=10)
        total_frame.pack(fill=tk.X)

        status_frame = ttk.Frame(total_frame)
        status_frame.pack(side=tk.LEFT)
        ttk.Label(status_frame, text="Order Status:").pack(side=tk.LEFT)
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(status_frame, textvariable=self.status_var, values=['Current', 'Sended'],
                                         state='readonly')
        self.status_combo.pack(side=tk.LEFT, padx=5)
        self.status_combo.bind('<<ComboboxSelected>>', self.on_status_change)

        self.lbl_expected_total = ttk.Label(total_frame, text="Total Expected: ₹0.00", font=("Helvetica", 12, "bold"),
                                            foreground="blue")
        self.lbl_expected_total.pack(side=tk.LEFT, padx=20)

        self.delete_item_button = ttk.Button(total_frame, text="Delete Selected", command=self.delete_selected_item)
        self.delete_item_button.pack(side=tk.RIGHT, padx=2)
        self.save_button = ttk.Button(total_frame, text="Save Order", command=self.save_order)
        self.save_button.pack(side=tk.RIGHT, padx=2)

    def trigger_scraper(self, event=None):
        part_no = self.item_part_no_var.get().strip()
        if not part_no: return
        self.lbl_scraper_status.config(text="Searching...", foreground="blue")
        threading.Thread(target=search_thread, args=(part_no,), daemon=True).start()

    def process_queue(self):
        try:
            while True:
                msg_type, payload = gui_queue.get_nowait()
                if msg_type == "status":
                    self.lbl_scraper_status.config(text=payload, foreground="black")
                elif msg_type == "light":
                    color = payload
                    fill = "#00ff00" if color == "green" else "#ffa500" if color == "orange" else "#ff0000"
                    self.canvas_light.itemconfig(self.light_circle, fill=fill)
                elif msg_type == "not_found":
                    self.lbl_scraper_status.config(text=payload, foreground="red")
                elif msg_type == "success":
                    self.lbl_scraper_status.config(text="Data Found!", foreground="green")
                    self.item_name_var.set(payload.get("Description", ""))
                    self.item_moq_var.set(payload.get("MOQ", "-"))
                    try:
                        clean = float(payload.get("Price", "0").replace('₹', '').replace(',', '').strip())
                        self.item_price_var.set(clean)
                    except:
                        self.item_price_var.set(0.0)
        except queue.Empty:
            pass
        self.after(100, self.process_queue)

    def add_item_to_order(self):
        name = self.item_name_var.get()
        part_no = self.item_part_no_var.get().strip()
        if self._check_duplicate_part_no(part_no): return
        try:
            qty = self.item_qty_var.get()
        except:
            qty = 0
        if not name or qty <= 0:
            messagebox.showwarning("Error", "Check inputs.")
            return

        try:
            price = self.item_price_var.get()
        except:
            price = 0.0
        try:
            dlp = self.item_dlp_var.get()
        except:
            dlp = 24.0
        moq = self.item_moq_var.get()

        self.items_in_order.append({
            'name': name, 'part_no': part_no, 'qty': qty,
            'price': price, 'vehicle': '', 'brand': '',
            'moq': moq, 'dlp': dlp
        })
        self.refresh_bill_treeview()
        self.clear_item_fields()
        self.part_no_entry.focus()

    def refresh_bill_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        grand_total = 0.0
        for i, item in enumerate(self.items_in_order):
            base_cost = item['price'] * item['qty']
            discount = base_cost * (item['dlp'] / 100.0)
            total = base_cost - discount
            grand_total += total

            self.tree.insert("", tk.END, values=(
                i + 1, item['part_no'], item['qty'], item['name'],
                format_currency(item['price']), item['moq'],
                f"{item['dlp']}%", format_currency(total)
            ))
        self.lbl_expected_total.config(text=f"Total Expected: {format_currency(grand_total)}")

    def clear_item_fields(self):
        self.item_name_var.set('')
        self.item_part_no_var.set('')
        self.item_qty_var.set(1)
        self.item_price_var.set(0.0)
        self.item_moq_var.set('')
        # Keep DLP at 24 for Hero
        self.item_dlp_var.set(24.0)
        self.part_no_entry.focus()


class NonOEMOrderPage(OrderPage):
    def __init__(self, parent, db, app):
        super().__init__(parent, db, app)
        self.current_party_id = self.db.get_or_create_id('parties', 'Non OEM')
        self.load_party_data_safe()

    def _setup_widgets(self):
        header = ttk.Frame(self, padding=10);
        header.pack(fill=tk.X)
        ttk.Button(header, text="Start New Order", command=self.start_new_order).pack(side=tk.LEFT)
        ttk.Label(header, text="  [Non-OEM Mode]", font=("Arial", 10, "bold"), foreground="green").pack(side=tk.LEFT)

        sel_frame = ttk.Frame(self, padding=10);
        sel_frame.pack(fill=tk.X)
        self.party_var = tk.StringVar(value="Non OEM")  # Dummy var to prevent crash

        of = ttk.LabelFrame(sel_frame, text="Select Non-OEM Order", padding=5)
        of.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        self.order_var = tk.StringVar()
        self.order_combo = ttk.Combobox(of, textvariable=self.order_var, state='readonly')
        self.order_combo.pack(fill=tk.X, expand=True)
        self.order_combo.bind('<<ComboboxSelected>>', self.on_order_selected)

        self._create_item_widgets();
        self._create_bill_details_widgets()

    def _create_item_widgets(self):
        self.item_frame = ttk.LabelFrame(self, text="Add Non-OEM Item", padding=10);
        self.item_frame.pack(fill=tk.X, padx=10, pady=5)
        for i in range(11): self.item_frame.columnconfigure(i, weight=1)

        self.item_name_var = tk.StringVar();
        self.item_part_no_var = tk.StringVar()
        self.vehicle_var = tk.StringVar();
        self.brand_var = tk.StringVar();
        self.item_qty_var = tk.IntVar(value=1)
        self.item_part_no_var.trace("w", self._force_uppercase)

        ttk.Label(self.item_frame, text="Item Name:").grid(row=0, column=0)
        self.item_name_entry = ttk.Entry(self.item_frame, textvariable=self.item_name_var)
        self.item_name_entry.grid(row=0, column=1, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Part No:").grid(row=0, column=2)
        self.part_no_entry = ttk.Entry(self.item_frame, textvariable=self.item_part_no_var)
        self.part_no_entry.grid(row=0, column=3, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Vehicle:").grid(row=0, column=4)
        self.vehicle_entry = ttk.Entry(self.item_frame, textvariable=self.vehicle_var)
        self.vehicle_entry.grid(row=0, column=5, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Brand:").grid(row=0, column=6)
        self.brand_entry = ttk.Entry(self.item_frame, textvariable=self.brand_var)
        self.brand_entry.grid(row=0, column=7, sticky=tk.EW, padx=2)

        ttk.Label(self.item_frame, text="Qty:").grid(row=0, column=8)
        self.qty_entry = ttk.Entry(self.item_frame, textvariable=self.item_qty_var, width=5)
        self.qty_entry.grid(row=0, column=9, sticky=tk.W, padx=2)

        self.add_item_button = ttk.Button(self.item_frame, text="Add", command=self.add_item_to_order)
        self.add_item_button.grid(row=0, column=10, padx=5)

        self.item_name_entry.bind("<Return>", lambda e: self.part_no_entry.focus())
        self.part_no_entry.bind("<Return>", lambda e: self.vehicle_entry.focus())
        self.vehicle_entry.bind("<Return>", lambda e: self.brand_entry.focus())
        self.brand_entry.bind("<Return>", lambda e: self.qty_entry.focus())
        self.qty_entry.bind("<Return>", lambda e: self.add_item_to_order())

    def _create_bill_details_widgets(self):
        self.bill_list_frame = ttk.LabelFrame(self, text="Current Bill Details", padding=10)
        self.bill_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ("sr_no", "item_name", "part_no", "vehicle", "brand", "qty")
        headers = {"sr_no": "SN", "item_name": "Item Name", "part_no": "Part No", "vehicle": "Vehicle",
                   "brand": "Brand", "qty": "Qty"}
        widths = {"sr_no": 40, "item_name": 200, "part_no": 100, "vehicle": 100, "brand": 100, "qty": 60}

        self.tree = ttk.Treeview(self.bill_list_frame, columns=cols, show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)
        for col in cols:
            self.tree.heading(col, text=headers[col]);
            self.tree.column(col, width=widths[col], anchor=tk.CENTER if col != "item_name" else tk.W)

        total_frame = ttk.Frame(self, padding=10);
        total_frame.pack(fill=tk.X)
        status_frame = ttk.Frame(total_frame);
        status_frame.pack(side=tk.LEFT)
        ttk.Label(status_frame, text="Order Status:").pack(side=tk.LEFT)
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(status_frame, textvariable=self.status_var, values=['Current', 'Sended'],
                                         state='readonly')
        self.status_combo.pack(side=tk.LEFT, padx=5);
        self.status_combo.bind('<<ComboboxSelected>>', self.on_status_change)

        self.lbl_expected_total = ttk.Label(total_frame, text="")  # Empty label to prevent crash
        self.lbl_expected_total.pack(side=tk.LEFT, padx=20)

        self.delete_item_button = ttk.Button(total_frame, text="Delete Selected", command=self.delete_selected_item)
        self.delete_item_button.pack(side=tk.RIGHT, padx=2)
        self.save_button = ttk.Button(total_frame, text="Save Order", command=self.save_order)
        self.save_button.pack(side=tk.RIGHT, padx=2)

    def load_party_data_safe(self):
        if self.current_party_id is None: return
        all_orders = self.db.get_all_orders_for_party(self.current_party_id)
        display_list = [f"{order[1]} ({order[2]})" for order in all_orders]
        current = next((order for order in all_orders if order[2] == 'Current'), None)
        if current:
            self.load_order_details(current[0], f"{current[1]} (Current)");
            self.order_var.set(f"{current[1]} (Current)")
        else:
            self.clear_bill_details();
            self.pending_order_number = self.db.generate_order_number()
            if hasattr(self, 'item_frame'): self.item_frame.config(text=f"Items for Order: {self.pending_order_number}")
            self.status_var.set("Current");
            self.on_status_change();
            self.order_var.set("New Order (Current)")
        self.order_combo['values'] = display_list

    def add_item_to_order(self):
        name = self.item_name_var.get();
        part = self.item_part_no_var.get().strip()
        if self._check_duplicate_part_no(part): return
        try:
            qty = self.item_qty_var.get()
        except:
            qty = 0
        if not name or qty <= 0: messagebox.showwarning("Error", "Check input"); return
        self.items_in_order.append(
            {'name': name, 'part_no': part, 'qty': qty, 'price': 0.0, 'vehicle': self.vehicle_var.get(),
             'brand': self.brand_var.get(), 'moq': '-', 'dlp': 0.0})
        self.refresh_bill_treeview();
        self.clear_item_fields();
        self.item_name_entry.focus()

    def refresh_bill_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for i, item in enumerate(self.items_in_order):
            self.tree.insert("", tk.END, values=(i + 1, item['name'], item['part_no'], item.get('vehicle', ''),
                                                 item.get('brand', ''), item['qty']))

    def clear_item_fields(self):
        self.item_name_var.set('');
        self.item_part_no_var.set('');
        self.vehicle_var.set('');
        self.brand_var.set('');
        self.item_qty_var.set(1);
        self.item_name_entry.focus()


class HistoryPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent); self.db = db; self.app = app
        controls = ttk.Frame(self, padding=10); controls.pack(fill=tk.X)
        ttk.Button(controls, text="Refresh", command=self.refresh_orders).pack(side=tk.LEFT)
        ttk.Button(controls, text="View/Edit Selected", command=self.edit_order).pack(side=tk.LEFT, padx=2)
        ttk.Button(controls, text="Delete Order", command=self.delete_order).pack(side=tk.LEFT, padx=2)
        if PDF_ENABLED: ttk.Button(controls, text="Export to PDF", command=self.export_to_pdf).pack(side=tk.LEFT, padx=2)
        else: ttk.Label(controls, text="PDF export disabled. Run 'pip install reportlab'").pack(side=tk.LEFT, padx=5)

        # --- NEW XLSX BUTTON ---
        if XLSX_ENABLED:
            ttk.Button(controls, text="Export to XLSX", command=self.export_to_xlsx).pack(side=tk.LEFT, padx=2)
        else:
            ttk.Label(controls, text="XLSX export disabled. Run 'pip install openpyxl'").pack(side=tk.LEFT, padx=5)

        # --- NEW SEARCH BAR ---
        search_frame = ttk.Frame(controls)
        search_frame.pack(side=tk.RIGHT, padx=5) # Pack to the right
        ttk.Label(search_frame, text="Search by Part No:").pack(side=tk.LEFT)
        self.part_no_search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.part_no_search_var, width=20).pack(side=tk.LEFT, padx=(5,2))
        ttk.Button(search_frame, text="Search", command=self.search_by_part_no).pack(side=tk.LEFT, padx=(2,0))
        # --- END NEW SEARCH BAR ---
        
        cols = ("order_no", "party_name", "creation_date", "last_saved", "status"); self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for col in cols: self.tree.heading(col, text=col.replace('_', ' ').title())
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.tree.tag_configure('Current', background='#F8D7DA', foreground='#721C24')  # Red
        self.tree.tag_configure('Sended', background='#D4EDDA', foreground='#155724')  # Green
        self.refresh_orders()

    def refresh_orders(self, data=None):
        if hasattr(self, 'part_no_search_var'):
            self.part_no_search_var.set("") # Clear search box

        for i in self.tree.get_children(): self.tree.delete(i)
        
        if data is None:
            order_list = self.db.get_all_orders()
        else:
            order_list = data

        for order in order_list:
            formatted = list(order)
            if formatted[3] is None: formatted[3] = formatted[2]  # Keep this logic

            # --- NEW: Determine row tag based on status ---
            # Status is at index 4 in the 'order' tuple
            status = order[4]
            row_tag = ''

            if status == 'Current':
                row_tag = 'Current'
            elif status == 'Sended':
                row_tag = 'Sended'
            # --- END NEW ---

            # Insert the row with the new tag
            self.tree.insert("", tk.END, values=formatted, tags=(row_tag,))

    def edit_order(self):
        if not self.tree.selection(): messagebox.showwarning("No Selection", "Please select an order."); return
        values = self.tree.item(self.tree.selection()[0])['values']
        order_number, party_name = values[0], values[1]
        if party_name == "Non OEM":
            self.app.notebook.select(self.app.nonoem_page); self.app.nonoem_page.load_order_from_history(party_name, order_number)
        else:
            self.app.notebook.select(self.app.order_page); self.app.order_page.load_order_from_history(party_name, order_number)

    def delete_order(self):
        """Deletes the selected order from the database."""
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select an order to delete.")
            return

        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to permanently delete this order and all its items?\nThis action cannot be undone."):
            return

        try:
            selected_item = self.tree.selection()[0]
            values = self.tree.item(selected_item)['values']
            order_number = values[0]
            
            order_id = self.db.get_order_id_by_number(order_number)
            if not order_id:
                messagebox.showerror("Error", f"Could not find order ID for '{order_number}'.")
                return

            self.db.delete_order(order_id)
            self.refresh_orders()
            messagebox.showinfo("Success", f"Order '{order_number}' has been deleted.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while deleting the order: {e}")

    def search_by_part_no(self):
        """Filters the order list by part number."""
        part_no = self.part_no_search_var.get().strip()
        if not part_no:
            self.refresh_orders() # Just refresh if empty
            return
        
        results = self.db.search_orders_by_part_number(part_no)
        if not results:
            messagebox.showinfo("No Results", f"No orders found containing part number '{part_no}'.")
        
        self.refresh_orders(results)

    def export_to_pdf(self):
        if not self.tree.selection(): messagebox.showwarning("No Selection",
                                                             "Please select an order to export."); return
        values = self.tree.item(self.tree.selection()[0])['values']
        order_number, party_name, date_str = values[0], values[1], values[2]
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            date_obj = datetime.now()

        filename_date = date_obj.strftime("%d-%m-%y")
        filename = f"{str(party_name).replace(' ', '_')}({filename_date}).pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")],
                                                initialfile=filename)
        if not filepath: return
        order_id = self.db.get_order_id_by_number(order_number)
        if not order_id: messagebox.showerror("Error", "Could not find order details."); return

        items = self.db.get_order_items(order_id)

        def draw_watermark(canvas, doc):
            """
            This function is called for every page.
            It draws the image on top of the content.
            """
            # --- Configuration ---
            image_path = resource_path("branding.png")  # <--- IMPORTANT: Change this

            # Check if the image file exists
            if not os.path.exists(image_path):
                print(f"WARNING: Watermark image not found at {image_path}")
                return  # Silently skip if no image

            canvas.saveState()

            # --- Sizing (3x3 inch as you requested) ---
            img_width = 1 * inch
            img_height = 1 * inch

            # --- Positioning (Top-Right) ---
            # Get page size (A4)
            page_width, page_height = A4

            # Set a 0.5-inch margin from the top-right edges
            margin = 0.5 * inch
            x = page_width - img_width - margin
            y = page_height - img_height - margin

            # --- Draw the Image ---
            #canvas.setFillAlpha(0.3)  # Set transparency (e.g., 30% opaque)
            canvas.drawImage(
                image_path,
                x,
                y,
                width=img_width,
                height=img_height,
                mask='auto',  # Handles PNG transparency
                preserveAspectRatio=True
            )

            canvas.restoreState()

        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4);
            styles = getSampleStyleSheet();
            styles['Normal'].fontName = 'Helvetica'
            styles['h1'].fontName = 'Helvetica-Bold'
            elements = []

            title_style = ParagraphStyle('TitleCustom', parent=styles['h1'], alignment=TA_CENTER)
            elements.extend([Paragraph("ORDER FORM", title_style), Spacer(1, 0.2 * inch),
                             Paragraph(f"<b>Party Name:</b> {escape(party_name)}", styles['Normal']),
                             Paragraph(f"<b>Order No:</b> {escape(order_number)}", styles['Normal']),
                             Paragraph(f"<b>Date:</b> {date_obj.strftime('%d-%B-%Y')}", styles['Normal']),
                             Spacer(1, 0.3 * inch)])

            is_nonoem = party_name == "Non OEM"
            if is_nonoem:
                data = [['SN', 'Item Name', 'Part No', 'Vehicle', 'Brand', 'Quantity']]
                # Unpack all 8, but only use relevant ones
                for i, (name, part_no, qty, price, vehicle, brand, moq, dlp) in enumerate(items, 1):
                    data.append(
                        [i, escape(name), escape(part_no) or 'N/A', escape(vehicle) or 'N/A', escape(brand) or 'N/A',
                         f"{qty} PCS"])
                table = Table(data,
                              colWidths=[0.4 * inch, 2.5 * inch, 1.5 * inch, 1.25 * inch, 1.25 * inch, 0.6 * inch],
                              repeatRows=1)
            else:
                data = [['SN', 'Item Name', 'Part No', 'Quantity']]
                # Unpack all 8, ignore price, moq, dlp
                for i, (name, part_no, qty, price, vehicle, brand, moq, dlp) in enumerate(items, 1):
                    data.append([i, escape(name), escape(part_no) or 'N/A', f"{qty} PCS"])
                table = Table(data, colWidths=[0.5 * inch, 4 * inch, 2 * inch, 1 * inch], repeatRows=1)

            style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)])
            table.setStyle(style)
            for i, _ in enumerate(data):
                if i % 2 == 0 and i > 0: style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#DCE6F1'))
            table.setStyle(style);
            elements.append(table);
            doc.build(elements)  # assuming watermark function is passed correctly
            messagebox.showinfo("Success", f"PDF successfully exported to\n{filepath}")
        except Exception as e: messagebox.showerror("PDF Export Error", f"An error occurred: {e}")

    def export_to_xlsx(self):
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select an order to export.")
            return

        # 1. Get selected order info
        values = self.tree.item(self.tree.selection()[0])['values']
        order_number, party_name, date_str = values[0], values[1], values[2]

        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            date_obj = datetime.now()

        # 2. Ask for save file path
        filename_date = date_obj.strftime("%d-%m-%y")
        filename = f"{str(party_name).replace(' ', '_')}({filename_date}).xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=filename
        )
        if not filepath:
            return

        # 3. Get order items from DB
        order_id = self.db.get_order_id_by_number(order_number)
        if not order_id:
            messagebox.showerror("Error", "Could not find order details.")
            return
        items = self.db.get_order_items(order_id)
        is_nonoem = party_name == "Non OEM"

        # 4. Create and populate workbook
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Order Details"

            # --- Define Styles ---
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left')
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # --- Title ---
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "ORDER FORM"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = center_align

            # --- Order Info ---
            ws['A3'] = "Party Name:"
            ws['A3'].font = bold_font
            ws['B3'] = party_name

            ws['A4'] = "Order No:"
            ws['A4'].font = bold_font
            ws['B4'] = order_number

            ws['A5'] = "Date:"
            ws['A5'].font = bold_font
            ws['B5'] = date_obj.strftime('%d-%B-%Y')

            # --- Items Table Headers ---
            if is_nonoem:
                headers = ['SN', 'Item Name', 'Part No', 'Vehicle', 'Brand', 'Quantity']
                col_widths = {'A': 5, 'B': 35, 'C': 20, 'D': 20, 'E': 20, 'F': 10}
            else:
                headers = ['SN', 'Item Name', 'Part No', 'Quantity']
                col_widths = {'A': 5, 'B': 45, 'C': 25, 'D': 10}

            # Start table at row 7
            ws.append(headers)
            header_row = ws.max_row

            for i, header in enumerate(headers):
                cell = ws.cell(row=header_row, column=i + 1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # --- Items Table Data ---
            for i, (name, part_no, qty, _, vehicle, brand) in enumerate(items, 1):
                # Use escape() to handle any problematic characters, though less critical for XLSX
                if is_nonoem:
                    row_data = [i, escape(name), escape(part_no) or 'N/A', escape(vehicle) or 'N/A',
                                escape(brand) or 'N/A', f"{qty} PCS"]
                else:
                    row_data = [i, escape(name), escape(part_no) or 'N/A', f"{qty} PCS"]
                ws.append(row_data)

            # --- Set Column Widths ---
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 5. Save the file
            wb.save(filepath)
            messagebox.showinfo("Success", f"Excel file successfully exported to\n{filepath}")

        except Exception as e:
            messagebox.showerror("XLSX Export Error", f"An error occurred while creating the Excel file: {e}")

class AccountingPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent); self.db = db; self.app = app; self.vars = {}
        self.editing_invoice_id = None
        self._setup_widgets()

    def _setup_widgets(self):
        # --- Company Management ---
        company_bar = ttk.LabelFrame(self, text="Companies", padding=5)
        company_bar.pack(fill=tk.X, padx=10, pady=5, expand=False)
        ttk.Button(company_bar, text="Add +", command=self.add_company).pack(side=tk.LEFT)
        self.company_buttons_frame = ttk.Frame(company_bar)
        self.company_buttons_frame.pack(side=tk.LEFT, fill=tk.X)

        # --- Invoice Form ---
        self.invoice_form = ttk.LabelFrame(self, text="Add/Edit Invoice", padding=10)
        form_fields = {
            "Company Name": {"readonly": True}, "KAR ID": {"readonly": True}, "CUSTOMER CODE IF ANY": {}, "ORDER NUMBER": {}, "INVOICE NUMBER": {},
            "AMOUNT": {}, "INVOICE DATE": {}, "PAYMENT DATE": {}, "PARTIAL PAYMENT DATE": {}, "DEBIT BANK NAME": {},
            "ACCOUNT NUMBER": {}, "TRANSACTION REF. NO/ CHEQ NO": {} # Removed Reference ID from here
            # "REFERENCE ID/UTR": {} # Moved below
        }
        row, col = 0, 0
        self.invoice_form.columnconfigure(1, weight=1)
        self.invoice_form.columnconfigure(3, weight=1)
        for label, options in form_fields.items():
            ttk.Label(self.invoice_form, text=label + ":").grid(row=row, column=col, sticky=tk.W, padx=5, pady=2)
            var = tk.StringVar()
            if "DATE" in label and CALENDAR_ENABLED:
                entry = DateEntry(self.invoice_form, textvariable=var, date_pattern='dd-mm-yyyy', width=17)
            else:
                entry = ttk.Entry(self.invoice_form, textvariable=var, state='readonly' if options.get("readonly") else 'normal')
            entry.grid(row=row, column=col+1, sticky=tk.EW, padx=5, pady=2)
            self.vars[label] = var
            col += 2
            # Start new row after every 2 fields (4 columns total)
            if col >= 4:
                 col = 0; row += 1

        # --- Payment Mode & Status on a separate row ---
        current_row = row # Remember the row we ended on
        if col != 0: # If the last field didn't complete a row, move to the next
            current_row += 1

        ttk.Label(self.invoice_form, text="PAYMENT MODE:").grid(row=current_row, column=0, sticky=tk.W, padx=5, pady=2)
        payment_frame = ttk.Frame(self.invoice_form)
        payment_frame.grid(row=current_row, column=1, sticky=tk.W)
        self.vars["PAYMENT MODE"] = {}
        for mode in ["CASH", "UPI", "IMPS","CHEQUE","NEFT"]:
            var = tk.BooleanVar()
            ttk.Checkbutton(payment_frame, text=mode, variable=var).pack(side=tk.LEFT)
            self.vars["PAYMENT MODE"][mode] = var

        ttk.Label(self.invoice_form, text="STATUS:").grid(row=current_row, column=2, sticky=tk.W, padx=5, pady=2)
        self.vars["STATUS"] = tk.StringVar()
        ttk.Combobox(self.invoice_form, textvariable=self.vars["STATUS"], values=["PAID", "UNPAID"], state='readonly').grid(row=current_row, column=3, sticky=tk.EW, padx=5, pady=2)

        # --- Reference ID on the next row ---
        next_row = current_row + 1
        ttk.Label(self.invoice_form, text="REFERENCE ID/UTR:").grid(row=next_row, column=0, sticky=tk.W, padx=5, pady=2)
        self.vars["REFERENCE ID/UTR"] = tk.StringVar()
        ttk.Entry(self.invoice_form, textvariable=self.vars["REFERENCE ID/UTR"]).grid(row=next_row, column=1, sticky=tk.EW, padx=5, pady=2)

        # --- Save/Clear Button Frame ---
        btn_frame = ttk.Frame(self.invoice_form)
        btn_frame.grid(row=next_row, column=3, sticky=tk.E, pady=10, padx=5)

        self.save_button = ttk.Button(btn_frame, text="Save Invoice", command=self.save_invoice)
        self.save_button.pack(side=tk.LEFT, padx=5)

        self.clear_button = ttk.Button(btn_frame, text="Clear", command=self.clear_invoice_form)
        self.clear_button.pack(side=tk.LEFT)

        # Initially hide the form using pack_forget
        self.invoice_form.pack(fill=tk.X, padx=10, pady=10, expand=False)
        self.invoice_form.pack_forget()

        # --- Invoice Table ---
        self.table_frame = ttk.LabelFrame(self, text="All Invoices", padding=10)
        self.table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # --- Search Bar (within table_frame) ---
        search_frame = ttk.Frame(self.table_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(search_frame, text="Search By:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_field_var = tk.StringVar()
        search_combo = ttk.Combobox(search_frame, textvariable=self.search_field_var, 
                                    values=["Invoice No", "Reference ID"], state='readonly', width=15)
        search_combo.pack(side=tk.LEFT, padx=5)
        search_combo.set("Invoice No") # Default value

        self.search_term_var = tk.StringVar() # New variable for the single search term
        ttk.Entry(search_frame, textvariable=self.search_term_var, width=30).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(search_frame, text="Search", command=self.perform_search).pack(side=tk.LEFT, padx=5) # Points to new method
        ttk.Button(search_frame, text="Refresh", command=lambda: self.refresh_invoices_table(None)).pack(side=tk.LEFT)
        ttk.Button(search_frame, text="Export to CSV", command=self.export_to_csv).pack(side=tk.LEFT, padx=5)
        # --- End Search Bar ---

        # Treeview and Scrollbars setup (parent is self.table_frame)
        cols = (
            "sn", "id", "company", "kar_id", "customer_code", "order_number", "invoice_number",
            "payment_mode", "amount", "status", "invoice_date", "payment_date",
            "partial_payment_date", "debit_bank_name", "account_number",
            "transaction_ref", "reference_id"
        )
        self.invoice_tree = ttk.Treeview(self.table_frame, columns=cols, show="headings")
        self.invoice_tree.heading("sn", text="SN")
        for col in cols[1:]: self.invoice_tree.heading(col, text=col.replace('_', ' ').title())
        self.invoice_tree.column("id", width=0, stretch=tk.NO)
        self.invoice_tree.column("sn", width=40, anchor=tk.CENTER)
        for col, w in {
            "company": 120, "kar_id": 60, "customer_code": 100, "order_number":100, "invoice_number": 100, "payment_mode": 120,
            "amount": 100, "status": 80, "invoice_date": 100, "payment_date": 100, "partial_payment_date": 120,
            "debit_bank_name": 120, "account_number": 120, "transaction_ref": 120, "reference_id": 120
        }.items(): self.invoice_tree.column(col, width=w, anchor=tk.W)
        x_scrollbar = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.invoice_tree.xview)
        x_scrollbar.pack(side=tk.BOTTOM, fill="x")
        y_scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.invoice_tree.yview)
        y_scrollbar.pack(side=tk.RIGHT, fill="y")
        self.invoice_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        self.invoice_tree.pack(fill=tk.BOTH, expand=True)

        self.invoice_tree.tag_configure('PAID', background='#D4EDDA', foreground='#155724')
        self.invoice_tree.tag_configure('UNPAID', background='#F8D7DA', foreground='#721C24')
        # --- Bottom Button Bar (Edit/Delete) ---
        bottom_btn_frame = ttk.Frame(self)
        bottom_btn_frame.pack(pady=5, expand=False)

        ttk.Button(bottom_btn_frame, text="Edit Selected Invoice", command=self.edit_selected_invoice).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(bottom_btn_frame, text="Delete Selected Invoice", command=self.delete_invoice).pack(side=tk.LEFT,padx=5)
        # Initial data load
        self.refresh_company_buttons(); self.refresh_invoices_table()

    def add_company(self):
        name = simpledialog.askstring("Add Company", "Enter Company Name:")
        if name and self.db.add_accounting_company(name): self.refresh_company_buttons()
        elif name: messagebox.showwarning("Warning", "Company already exists.")

    def delete_company(self, company_id, name):
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{name}' and all its invoices?"):
            # Check if the form for this company is currently shown
            if hasattr(self, 'selected_company_id') and self.selected_company_id == company_id:
                self.invoice_form.pack_forget() # Hide the form
                delattr(self, 'selected_company_id') # Remove the attribute
                
            self.db.delete_accounting_company(company_id)
            self.refresh_company_buttons()
            self.refresh_invoices_table()
            # self.invoice_form.pack_forget() # Removed redundant hide

    def refresh_company_buttons(self):
        for widget in self.company_buttons_frame.winfo_children(): widget.destroy()
        for company_id, name in self.db.get_accounting_companies():
            frame = ttk.Frame(self.company_buttons_frame)
            ttk.Button(frame, text=name, command=lambda cid=company_id, cname=name: self.show_invoice_form(cid, cname)).pack(side=tk.LEFT)
            ttk.Button(frame, text="x", width=2, command=lambda cid=company_id, cname=name: self.delete_company(cid, cname)).pack(side=tk.LEFT, padx=(0, 5))
            frame.pack(side=tk.LEFT)

    def show_invoice_form(self, company_id, name):
        # Simply pack the form again if it's hidden. Use 'before=self.table_frame'.
        if not self.invoice_form.winfo_ismapped():
             self.invoice_form.pack(fill=tk.X, padx=10, pady=10, expand=False, before=self.table_frame) # Use before=self.table_frame

        self.clear_invoice_form()
        self.vars["Company Name"].set(name)
        self.vars["KAR ID"].set(self.db.generate_kar_id())
        self.selected_company_id = company_id
        self.invoice_form.config(text=f"Add/Edit Invoice for {name}")

    def save_invoice(self):
        invoice_number = self.vars["INVOICE NUMBER"].get().strip()
        if not invoice_number:
            messagebox.showerror("Input Error", "Invoice Number cannot be empty.")
            return

        # Check for duplicates, ignoring the current ID if we are editing
        if self.db.check_invoice_number_exists(invoice_number, self.editing_invoice_id):
            messagebox.showwarning("Duplicate Entry", f"Invoice Number '{invoice_number}' already exists.")
            return

        payment_modes = [mode for mode, var in self.vars["PAYMENT MODE"].items() if var.get()]

        try:
            # Use safe_float to handle currency symbols or commas
            amount = safe_float(self.vars["AMOUNT"].get())
        except (ValueError, tk.TclError):
            messagebox.showerror("Input Error", "Amount must be a number.");
            return

        data = {
            # Use self.selected_company_id, which is set when opening the form or editing
            "company_id": self.selected_company_id,
            "kar_id": self.vars["KAR ID"].get(),
            "customer_code": self.vars["CUSTOMER CODE IF ANY"].get(),
            "order_number": self.vars["ORDER NUMBER"].get(),
            "invoice_number": invoice_number,
            "payment_mode": " + ".join(payment_modes),
            "amount": amount,
            "status": self.vars["STATUS"].get(),
            "invoice_date": self.vars["INVOICE DATE"].get(),
            "payment_date": self.vars["PAYMENT DATE"].get(),
            "partial_payment_date": self.vars["PARTIAL PAYMENT DATE"].get(),
            "debit_bank_name": self.vars["DEBIT BANK NAME"].get(),
            "account_number": self.vars["ACCOUNT NUMBER"].get(),
            "transaction_ref": self.vars["TRANSACTION REF. NO/ CHEQ NO"].get(),
            "reference_id": self.vars["REFERENCE ID/UTR"].get()
        }

        try:
            if self.editing_invoice_id:
                # --- This is an UPDATE ---
                self.db.update_invoice(self.editing_invoice_id, data)
                messagebox.showinfo("Success", "Invoice updated successfully.")
            else:
                # --- This is a CREATE NEW ---
                self.db.add_invoice(data)
                messagebox.showinfo("Success", "Invoice saved successfully.")

            self.refresh_invoices_table()

            # Clear the form and reset to "Add" mode
            company_name = self.vars["Company Name"].get()  # Remember company
            self.clear_invoice_form()
            self.vars["Company Name"].set(company_name)  # Restore company
            self.vars["KAR ID"].set(self.db.generate_kar_id())  # Get next ID

            # Hide the form after save/update
            self.invoice_form.pack_forget()

        except Exception as e:
            messagebox.showerror("Database Error", f"An error occurred: {e}")

    def delete_invoice(self):
        if not self.invoice_tree.selection(): messagebox.showwarning("No Selection", "Please select an invoice to delete."); return
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this invoice?"):
            selected_item = self.invoice_tree.selection()[0]
            invoice_id = self.invoice_tree.item(selected_item, 'values')[1]
            self.db.delete_invoice(invoice_id); self.refresh_invoices_table()

    def clear_invoice_form(self):
        """Clears all fields and resets the form to 'new invoice' mode."""
        self.editing_invoice_id = None  # Reset editing ID

        for key, var in self.vars.items():
            if key == "PAYMENT MODE":
                for check_var in var.values(): check_var.set(False)
            elif key not in ["Company Name", "KAR ID"]:
                var.set('')

        self.vars["STATUS"].set("UNPAID")

        # Reset KAR ID (but Company Name stays from selection)
        if hasattr(self, 'selected_company_id'):
            self.vars["KAR ID"].set(self.db.generate_kar_id())
        else:
            self.vars["KAR ID"].set("")
            self.vars["Company Name"].set("")

        # Reset button text and form title
        self.save_button.config(text="Save Invoice")
        if hasattr(self, 'selected_company_id'):
            company_name = self.vars["Company Name"].get()
            self.invoice_form.config(text=f"Add/Edit Invoice for {company_name}")
        else:
            self.invoice_form.config(text="Add/Edit Invoice")

    def edit_selected_invoice(self):
        """Loads the selected invoice's data into the form for editing."""
        if not self.invoice_tree.selection():
            messagebox.showwarning("No Selection", "Please select an invoice to edit.")
            return

        selected_item = self.invoice_tree.selection()[0]
        # Get the database ID (column index 1)
        invoice_id = self.invoice_tree.item(selected_item, 'values')[1]

        # Fetch all details from the database
        details = self.db.get_invoice_details_by_id(invoice_id)
        if not details:
            messagebox.showerror("Error", "Could not fetch invoice details.")
            return

        # --- Load data into the form ---
        self.clear_invoice_form()  # Clear it first

        self.editing_invoice_id = invoice_id
        self.selected_company_id = details['company_id']  # Set the company context

        # Populate text/combo vars
        self.vars["Company Name"].set(details['company_name'])
        self.vars["KAR ID"].set(details['kar_id'])
        self.vars["CUSTOMER CODE IF ANY"].set(details['customer_code'])
        self.vars["ORDER NUMBER"].set(details['order_number'])
        self.vars["INVOICE NUMBER"].set(details['invoice_number'])
        self.vars["AMOUNT"].set(format_currency(details['amount']))  # Use format_currency
        self.vars["STATUS"].set(details['status'])
        self.vars["INVOICE DATE"].set(details['invoice_date'])
        self.vars["PAYMENT DATE"].set(details['payment_date'])
        self.vars["PARTIAL PAYMENT DATE"].set(details['partial_payment_date'])
        self.vars["DEBIT BANK NAME"].set(details['debit_bank_name'])
        self.vars["ACCOUNT NUMBER"].set(details['account_number'])
        self.vars["TRANSACTION REF. NO/ CHEQ NO"].set(details['transaction_ref'])
        self.vars["REFERENCE ID/UTR"].set(details['reference_id'])

        # Populate payment mode checkboxes
        payment_modes = details['payment_mode'].split(' + ')
        for mode in payment_modes:
            if mode in self.vars["PAYMENT MODE"]:
                self.vars["PAYMENT MODE"][mode].set(True)

        # Update button text and form title
        self.save_button.config(text="Update Invoice")
        self.invoice_form.config(text=f"Editing Invoice: {details['invoice_number']}")

        # Show the form if it was hidden
        if not self.invoice_form.winfo_ismapped():
            self.invoice_form.pack(fill=tk.X, padx=10, pady=10, expand=False, before=self.table_frame)

    def refresh_invoices_table(self, data=None):
        """Refreshes the invoice table with optional filtered data."""
        # Clear search bar if refreshing all
        if data is None:
            self.search_term_var.set("") # Use the new search variable
            
        for i in self.invoice_tree.get_children(): self.invoice_tree.delete(i)
        
        if data is None:
            data = self.db.get_all_invoices()

        for i, row in enumerate(data, 1):
            formatted_row = list(row)
            formatted_row[7] = format_currency(row[7])  # Format amount

            # --- NEW: Determine row tag based on status ---
            # The status is at index 8 in the raw 'row' from the DB
            status = row[8]
            row_tag = ''

            if status == 'PAID':
                row_tag = 'PAID'
            elif status == 'UNPAID':
                row_tag = 'UNPAID'
            # --- END NEW ---

            # Insert the row with the new tag
            self.invoice_tree.insert("", "end", values=(i, *formatted_row), tags=(row_tag,))

    def export_to_csv(self):
        """Exports the current data in the invoice_tree to a CSV file."""
        
        # 1. Get all item IDs from the tree
        item_ids = self.invoice_tree.get_children()
        
        if not item_ids:
            messagebox.showinfo("No Data", "There is no data in the table to export.")
            return

        # 2. Ask user for a save location
        try:
            now = datetime.now().strftime("%Y-%m-%d_%H%M")
            filename = f"accounting_export_{now}.csv"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if not filepath:
                return # User cancelled

            # 3. Get headers from the Treeview
            cols = self.invoice_tree['columns']
            headers = [self.invoice_tree.heading(col, 'text') for col in cols]
            
            # 4. Write data to the CSV file
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write the headers
                writer.writerow(headers)
                
                # Write the data rows
                for item_id in item_ids:
                    values = self.invoice_tree.item(item_id, 'values')
                    writer.writerow(values)
                    
            messagebox.showinfo("Success", f"Data exported successfully to\n{filepath}")

        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred while exporting:\n{e}")

    def perform_search(self):
        """Filters the invoice table based on the selected field and search query."""
        search_term = self.search_term_var.get().strip()
        search_field = self.search_field_var.get()

        if not search_term:
            self.refresh_invoices_table(None)
            return
            
        search_results = []
        if search_field == "Invoice No":
            search_results = self.db.search_invoices_by_number(search_term)
        elif search_field == "Reference ID":
            # --- IMPORTANT ---
            # You must implement 'search_invoices_by_ref' in your database class.
            # It should search both 'transaction_ref' and 'reference_id' columns.
            #
            # Example SQL for your db class:
            # query = """SELECT <your_columns> FROM accounting_invoices 
            #            WHERE (transaction_ref LIKE ? OR reference_id LIKE ?)"""
            # wildcard_term = f'%{search_term}%'
            # self.cursor.execute(query, (wildcard_term, wildcard_term))
            # return self.cursor.fetchall()
            #
            search_results = self.db.search_invoices_by_ref(search_term)
        
        if not search_results:
            messagebox.showinfo("No Results", f"No invoices found matching '{search_term}' for {search_field}.")
        
        self.refresh_invoices_table(search_results)

class PartRequestPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent); self.db = db; self.app = app; self.vars = {}; self.editing_id = None
        self._setup_widgets()

    def _setup_widgets(self):
        # Main layout frames
        form_container = ttk.Frame(self); form_container.pack(fill=tk.X, padx=10, pady=5)
        table_container = ttk.Frame(self); table_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # --- Form Widgets ---
        form = ttk.LabelFrame(form_container, text="Part Request Details", padding=10); form.pack(fill=tk.X)
        form.columnconfigure(1, weight=1); form.columnconfigure(3, weight=1); form.columnconfigure(5, weight=1)

        # Row 0: Request ID, Customer Name, Phone
        ttk.Label(form, text="Order ID:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.vars['request_id'] = tk.StringVar(value=self.db.generate_request_id())
        ttk.Entry(form, textvariable=self.vars['request_id'], state='readonly').grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)
        
        ttk.Label(form, text="Customer Name:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        self.vars['customer_name'] = tk.StringVar()
        ttk.Entry(form, textvariable=self.vars['customer_name']).grid(row=0, column=3, sticky=tk.EW, padx=5, pady=2)
        
        ttk.Label(form, text="Phone Number:").grid(row=0, column=4, sticky=tk.W, padx=5, pady=2)
        self.vars['phone_number'] = tk.StringVar()
        ttk.Entry(form, textvariable=self.vars['phone_number']).grid(row=0, column=5, sticky=tk.EW, padx=5, pady=2)

        # Row 1: Security Amount, Payment Type, Request Date
        ttk.Label(form, text="Security Amount:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.vars['security_amount'] = tk.DoubleVar(value=0.0)
        ttk.Entry(form, textvariable=self.vars['security_amount']).grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)
        
        ttk.Label(form, text="Payment Type:").grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)
        self.vars['payment_type'] = tk.StringVar()
        ttk.Entry(form, textvariable=self.vars['payment_type']).grid(row=1, column=3, sticky=tk.EW, padx=5, pady=2) # Changed to Entry
        
        ttk.Label(form, text="Request Date:").grid(row=1, column=4, sticky=tk.W, padx=5, pady=2)
        self.vars['request_date'] = tk.StringVar()
        if CALENDAR_ENABLED:
            DateEntry(form, textvariable=self.vars['request_date'], date_pattern='dd-mm-yyyy').grid(row=1, column=5, sticky=tk.EW, padx=5, pady=2)
            self.vars['request_date'].set(datetime.now().strftime('%d-%m-%Y'))
        else:
            ttk.Entry(form, textvariable=self.vars['request_date']).grid(row=1, column=5, sticky=tk.EW, padx=5, pady=2)
            self.vars['request_date'].set(datetime.now().strftime('%d-%m-%Y'))

        # Row 2: Part Details (Text Widget)
        ttk.Label(form, text="Part Name/Number:").grid(row=2, column=0, sticky=tk.NW, padx=5, pady=5)
        self.part_details_text = tk.Text(form, height=4, width=40)
        self.part_details_text.grid(row=2, column=1, columnspan=5, sticky=tk.EW, padx=5, pady=2)

        # Row 3: Status
        ttk.Label(form, text="Status:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.vars['status'] = tk.StringVar()
        status_combo = ttk.Combobox(form, textvariable=self.vars['status'], values=["Pending", "Arrived", "Delivered"], state='readonly')
        status_combo.grid(row=3, column=1, sticky=tk.EW, padx=5, pady=2)
        self.vars['status'].set("Pending")

        # Row 4: Form Buttons (Clear, Save)
        button_frame = ttk.Frame(form)
        button_frame.grid(row=4, column=0, columnspan=6, sticky=tk.E, pady=10)
        ttk.Button(button_frame, text="Clear Form", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Save Request", command=self.save_request).pack(side=tk.LEFT, padx=5)


        # --- Table Widgets (Preview/Management Section) ---
        preview_frame = ttk.LabelFrame(table_container, text="All Part Requests", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True)

        # Actions Frame (Search, Refresh, Edit, Delete, Download)
        actions_frame = ttk.Frame(preview_frame); actions_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Search Widgets (Left side)
        ttk.Label(actions_frame, text="Search by ID:").pack(side=tk.LEFT, padx=(0, 2))
        self.search_id_var = tk.StringVar()
        ttk.Entry(actions_frame, textvariable=self.search_id_var, width=15).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(actions_frame, text="Search by Phone:").pack(side=tk.LEFT, padx=(0, 2))
        self.search_phone_var = tk.StringVar()
        ttk.Entry(actions_frame, textvariable=self.search_phone_var, width=15).pack(side=tk.LEFT)
        
        ttk.Button(actions_frame, text="Search", command=self.search_requests).pack(side=tk.LEFT, padx=5)
        ttk.Button(actions_frame, text="Refresh", command=lambda: self.refresh_table(None)).pack(side=tk.LEFT)
        
        # Management Buttons (Right side)
        button_group_right = ttk.Frame(actions_frame)
        button_group_right.pack(side=tk.RIGHT)
        
        if PDF_ENABLED:
            ttk.Button(button_group_right, text="Download Receipt", command=self.download_receipt).pack(side=tk.LEFT, padx=5)
        else:
            ttk.Label(button_group_right, text="Install 'reportlab' for PDF").pack(side=tk.LEFT, padx=5)
            
        ttk.Button(button_group_right, text="Delete Request", command=self.delete_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_group_right, text="Edit Request", command=self.edit_request).pack(side=tk.LEFT)

        # Treeview (Table) setup
        table_inner_frame = ttk.Frame(preview_frame) # Frame to hold treeview and scrollbars
        table_inner_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "request_id", "customer_name", "phone_number", "security_amount", "payment_type", "part_details", "request_date", "status")
        self.tree = ttk.Treeview(table_inner_frame, columns=cols, show="headings", height=10)
        
        # Scrollbars
        ysb = ttk.Scrollbar(table_inner_frame, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(table_inner_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        # Pack Treeview and Scrollbars
        xsb.pack(side=tk.BOTTOM, fill=tk.X)
        ysb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configure column headings and widths
        self.tree.column("id", width=0, stretch=tk.NO) # Hide internal DB ID
        self.tree.heading("request_id", text="Order ID"); self.tree.column("request_id", width=80, anchor=tk.W)
        self.tree.heading("customer_name", text="Customer Name"); self.tree.column("customer_name", width=150, anchor=tk.W)
        self.tree.heading("phone_number", text="Phone Number"); self.tree.column("phone_number", width=100, anchor=tk.W)
        self.tree.heading("security_amount", text="Security Amt"); self.tree.column("security_amount", width=100, anchor=tk.E) # Right align amount
        self.tree.heading("payment_type", text="Payment Type"); self.tree.column("payment_type", width=80, anchor=tk.W)
        self.tree.heading("part_details", text="Part Details"); self.tree.column("part_details", width=300, anchor=tk.W)
        self.tree.heading("request_date", text="Request Date"); self.tree.column("request_date", width=100, anchor=tk.CENTER)
        self.tree.heading("status", text="Status"); self.tree.column("status", width=80, anchor=tk.W)
        # --- NEW: Configure row tags for status colors ---
        self.tree.tag_configure('Pending', background='#FFC107', foreground='black')
        self.tree.tag_configure('Arrived', background='#17A2B8', foreground='black')
        self.tree.tag_configure('Delivered', background='#28A745', foreground='black')
        # --- END NEW ---
        # Initial data load
        self.refresh_table()

    def clear_form(self):
        """Clears all entry fields in the part request form."""
        self.editing_id = None # Reset editing state
        self.vars['request_id'].set(self.db.generate_request_id()) # Get new ID
        self.vars['customer_name'].set('')
        self.vars['phone_number'].set('')
        self.vars['security_amount'].set(0.0)
        self.vars['payment_type'].set('')
        self.part_details_text.delete('1.0', tk.END) # Clear Text widget
        self.vars['request_date'].set(datetime.now().strftime('%d-%m-%Y')) # Set current date
        self.vars['status'].set('Pending') # Default status

    def save_request(self):
        """Saves a new or updates an existing part request."""
        try:
            amount = self.vars['security_amount'].get()
        except tk.TclError:
            messagebox.showerror("Input Error", "Security Amount must be a valid number.")
            return

        # Prepare data dictionary from form variables
        data = {
            "customer_name": self.vars['customer_name'].get().strip(), 
            "phone_number": self.vars['phone_number'].get().strip(), 
            "security_amount": amount,
            "payment_type": self.vars['payment_type'].get().strip(), 
            "part_details": self.part_details_text.get('1.0', tk.END).strip(),
            "request_date": self.vars['request_date'].get(), 
            "status": self.vars['status'].get()
        }

        # Basic validation
        if not data["customer_name"] or not data["phone_number"]:
            messagebox.showerror("Input Error", "Customer Name and Phone Number are required.")
            return

        try:
            if self.editing_id:
                # Update existing record
                self.db.update_part_request(self.editing_id, data)
                messagebox.showinfo("Success", "Part request updated successfully.")
            else:
                # Add new record, including the generated request_id
                data_with_id = {'request_id': self.vars['request_id'].get(), **data}
                self.db.add_part_request(data_with_id)
                messagebox.showinfo("Success", "Part request saved successfully.")

            self.clear_form() # Clear form after successful save/update
            self.refresh_table() # Refresh the table view

        except Exception as e:
            messagebox.showerror("Database Error", f"An error occurred: {e}")


    def refresh_table(self, data=None):
        """Refreshes the Treeview with part request data."""
        # Clear search fields
        self.search_id_var.set("")
        self.search_phone_var.set("")
        
        # Clear existing table rows
        for i in self.tree.get_children():
            self.tree.delete(i)
            
        # Fetch data if not provided (e.g., on initial load or full refresh)
        if data is None:
            data = self.db.get_all_part_requests()
            
        # Populate table
        for row in data:
            formatted = list(row)
            formatted[4] = format_currency(row[4]) # Format security amount
            status = row[8]
            row_tag = ''  # Default tag (no color)
            if status == 'Pending':
                row_tag = 'Pending'
            elif status == 'Arrived':
                row_tag = 'Arrived'
            elif status == 'Delivered':
                row_tag = 'Delivered'

            self.tree.insert("", tk.END, values=formatted, tags=(row_tag,))

    def edit_request(self):
        """Loads the selected request from the table into the form for editing."""
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a request to edit.")
            return

        selected_item = self.tree.selection()[0]
        values = self.tree.item(selected_item)['values']

        try:
            # Unpack values based on the table column order
            db_id, request_id, customer_name, phone, amount_str, p_type, details, req_date, status = values
        except ValueError:
            messagebox.showerror("Error", "Could not read the selected row data.")
            return

        # Set the editing ID state
        self.editing_id = db_id

        # Populate form fields
        self.vars['request_id'].set(request_id) # Display the ID being edited
        self.vars['customer_name'].set(customer_name)
        self.vars['phone_number'].set(phone)
        # Use safe_float to convert formatted currency string back to float
        self.vars['security_amount'].set(safe_float(amount_str)) 
        self.vars['payment_type'].set(p_type)
        self.part_details_text.delete('1.0', tk.END) # Clear and insert details
        self.part_details_text.insert('1.0', details)
        self.vars['request_date'].set(req_date)
        self.vars['status'].set(status)
        
        # Optional: Bring the form into view if it's scrollable/hidden
        # form_container.focus_set() # Or scroll if needed

    def delete_request(self):
        """Deletes the selected part request from the database."""
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a request to delete.")
            return
            
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to permanently delete this part request?"):
            try:
                selected_item = self.tree.selection()[0]
                db_id = self.tree.item(selected_item, 'values')[0] # Get internal ID
                self.db.delete_part_request(db_id)
                self.refresh_table() # Refresh table after deletion
                messagebox.showinfo("Success", "Part request deleted.")
                # If the deleted item was being edited, clear the form
                if self.editing_id == db_id:
                     self.clear_form()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete request: {e}")

    def search_requests(self):
        """Searches for requests based on ID or Phone and refreshes the table."""
        search_id = self.search_id_var.get().strip()
        search_phone = self.search_phone_var.get().strip()
        
        results = self.db.search_part_requests(order_id=search_id, phone=search_phone)
        
        if not results:
             messagebox.showinfo("No Results", "No part requests found matching the criteria.")
             
        self.refresh_table(results) # Refresh table with search results

    def download_receipt(self):
        
        """Generates and prompts to save a PDF receipt for the selected request."""
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a request to download.")
            return
        if not PDF_ENABLED:
            messagebox.showerror("Error", "PDF library (reportlab) not found. Run 'pip install reportlab'.")
            return

        selected_item = self.tree.selection()[0]
        values = self.tree.item(selected_item)['values']
        def draw_watermark(canvas, doc):
            """
            This function is called for every page.
            It draws the image on top of the content.
            """
            # --- Configuration ---
            image_path = resource_path("branding.png")  # <--- IMPORTANT: Change this

            # Check if the image file exists
            if not os.path.exists(image_path):
                print(f"WARNING: Watermark image not found at {image_path}")
                return  # Silently skip if no image

            canvas.saveState()

            # --- Sizing (3x3 inch as you requested) ---
            img_width = 1 * inch
            img_height = 1 * inch

            # --- Positioning (Top-Right) ---
            # Get page size (A4)
            page_width, page_height = A4

            # Set a 0.5-inch margin from the top-right edges
            margin = 0.5 * inch
            x = page_width - img_width - margin
            y = page_height - img_height - margin

            # --- Draw the Image ---
            #canvas.setFillAlpha(0.3)  # Set transparency (e.g., 30% opaque)
            canvas.drawImage(
                image_path,
                x,
                y,
                width=img_width,
                height=img_height,
                mask='auto',  # Handles PNG transparency
                preserveAspectRatio=True
            )

            canvas.restoreState()
        try:
            db_id, request_id, customer_name, phone, amount_str, p_type, details, req_date, status = values
        except ValueError:
            messagebox.showerror("Error", "Could not read selected row data for PDF generation.")
            return

        # Prepare filename and ask for save location
        filename = f"Request_{request_id}_{str(customer_name).replace(' ', '_')}.pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")], initialfile=filename)
        if not filepath: return # User cancelled save dialog

        def draw_watermark(canvas, doc):
            """
            This function is called for every page.
            It draws the image on top of the content.
            """
            # --- Configuration ---
            image_path = resource_path("branding.png")  # <--- IMPORTANT: Change this

            # Check if the image file exists
            if not os.path.exists(image_path):
                print(f"WARNING: Watermark image not found at {image_path}")
                return  # Silently skip if no image

            canvas.saveState()

            # --- Sizing (3x3 inch as you requested) ---
            img_width = 1 * inch
            img_height = 1 * inch

            # --- Positioning (Top-Right) ---
            # Get page size (A4)
            page_width, page_height = A4

            # Set a 0.5-inch margin from the top-right edges
            margin = 0.5 * inch
            x = page_width - img_width - margin
            y = page_height - img_height - margin

            # --- Draw the Image ---
            #canvas.setFillAlpha(0.3)  # Set transparency (e.g., 30% opaque)
            canvas.drawImage(
                image_path,
                x,
                y,
                width=img_width,
                height=img_height,
                mask='auto',  # Handles PNG transparency
                preserveAspectRatio=True
            )

            canvas.restoreState()
        try:
            # --- PDF Generation Logic ---
            doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            styles = getSampleStyleSheet(); elements = []

            # Title
            title_style = ParagraphStyle('TitleCustom', parent=styles['h1'], alignment=TA_CENTER)
            elements.append(Paragraph("ON DEMAND ORDER", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Header Table (Customer Info)
            header_data = [
                ['Customer Name:', escape(str(customer_name)), 'Order ID:', escape(str(request_id))],
                ['Mob Number:', escape(str(phone)), 'Payment Type:', escape(str(p_type))],
            ]
            header_table = Table(header_data, colWidths=[1.2*inch, 2.8*inch, 1.2*inch, 2.3*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,-1), 'LEFT'), ('ALIGN', (2,0), (2,-1), 'LEFT'), # Labels left
                ('ALIGN', (1,0), (1,-1), 'LEFT'), ('ALIGN', (3,0), (3,-1), 'LEFT'), # Values left
                ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), # Bold labels
                ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'), # Bold labels
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.2*inch))
            
            # Format amount string correctly
            amount_float = safe_float(amount_str) # Convert formatted string back to float
            amount_str_formatted = format_currency(amount_float) # Re-format consistently
            
            # Escape details and handle newlines for Paragraph
            safe_details = escape(str(details)).replace('\n', '<br/>')

            # Main Content Table (Description, Amount)
            data = [
                ['DESCRIPTION', 'AMOUNT'],
                [Paragraph(safe_details, styles['Normal']), Paragraph(amount_str_formatted, styles['BodyText'])] # Use BodyText for potentially smaller font
            ]
            
            # Define table with adjusted row height for details potentially spanning multiple lines
            table = Table(data, colWidths=[5.5*inch, 2.0*inch], rowHeights=[None, None]) # Let ReportLab decide row height
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D3D3D3')), # Header background
                ('TEXTCOLOR',(0,0),(-1,0), colors.black),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), # Bold header
                ('ALIGN', (0,0), (-1,0), 'CENTER'), # Center header text
                ('ALIGN', (-1,1), (-1,1), 'RIGHT'), # Amount right-aligned
                ('VALIGN', (-1,1), (-1,1), 'TOP'), # Amount top-aligned
                ('ALIGN', (0,1), (0,1), 'LEFT'), # Description left-aligned
                ('VALIGN', (0,1), (0,1), 'TOP'), # Description top-aligned
                ('GRID', (0,0), (-1,-1), 1, colors.black), # Grid lines
                ('TOPPADDING', (0,1), (-1,-1), 6), # Padding within cells
                ('LEFTPADDING', (0,1), (-1,-1), 6),
                ('RIGHTPADDING', (0,1), (-1,-1), 6),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.5*inch))
            
            # Footer Table (Date, Signature)
            footer_data = [
                [f"DATE: {escape(str(req_date))}", 'Authorized Signatory']
            ]
            footer_table = Table(footer_data, colWidths=[3.75*inch, 3.75*inch])
            footer_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,0), 'LEFT'), # Date left
                ('ALIGN', (1,0), (1,0), 'RIGHT'), # Signature right
                ('FONTNAME', (1,0), (1,0), 'Helvetica-Bold'), # Bold signature label
                ('TOPPADDING', (0,0), (-1,-1), 20), # Space above footer
            ]))
            elements.append(footer_table)

            # Build the PDF document
            doc.build(elements, onFirstPage=draw_watermark, onLaterPages=draw_watermark)
            messagebox.showinfo("Success", "Receipt downloaded successfully.")
            
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"Failed to generate PDF: {e}")
            
# --- NEW: Dashboard Page ---
# --- NEW: Dashboard Page ---
class DashboardPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent)
        self.db = db
        self.app = app

        # Map month names to numbers
        self.month_map = {datetime(2000, m, 1).strftime('%B'): f"{m:02d}" for m in range(1, 13)}

        self._setup_styles()
        self._setup_widgets()
        self.update_summary()  # Load initial data

    def _setup_styles(self):
        """Create custom styles for our labels."""
        style = ttk.Style()
        style.configure("Dash.TLabel", font=("Helvetica", 16))
        style.configure("Paid.TLabel", foreground="#008800", font=("Helvetica", 18, "bold"))
        style.configure("Unpaid.TLabel", foreground="#CC0000", font=("Helvetica", 18, "bold"))
        style.configure("Total.TLabel", font=("Helvetica", 18, "bold"))

    def _setup_widgets(self):
        # --- Main Summary Frame ---
        summary_frame = ttk.LabelFrame(self, text="Monthly Accounting Summary", padding=20)
        summary_frame.pack(expand=False, fill=tk.X, padx=20, pady=20, anchor=tk.N)

        # --- Controls Frame ---
        controls_frame = ttk.Frame(summary_frame)
        controls_frame.pack(fill=tk.X, pady=(0, 20))
        controls_frame.columnconfigure(1, weight=1)
        controls_frame.columnconfigure(3, weight=1)

        # Month Selector
        ttk.Label(controls_frame, text="Month:").grid(row=0, column=0, padx=5, sticky=tk.W)
        self.month_var = tk.StringVar()
        month_combo = ttk.Combobox(controls_frame, textvariable=self.month_var,
                                   values=list(self.month_map.keys()), state='readonly')
        month_combo.grid(row=0, column=1, sticky=tk.EW, padx=5)
        month_combo.set(datetime.now().strftime('%B'))  # Default to current month
        month_combo.bind('<<ComboboxSelected>>', self.on_selection_change)

        # Year Selector
        ttk.Label(controls_frame, text="Year:").grid(row=0, column=2, padx=(20, 5), sticky=tk.W)
        self.year_var = tk.StringVar()
        current_year = datetime.now().year

        # --- MODIFIED RANGE: Starts from Last Year (2025) ---
        # range(start, stop) -> start is inclusive, stop is exclusive
        # This gives: [current_year - 1, current_year, ... +4 more]
        years = [str(y) for y in range(current_year - 1, current_year + 6)]

        year_combo = ttk.Combobox(controls_frame, textvariable=self.year_var, values=years, state='readonly')
        year_combo.grid(row=0, column=3, sticky=tk.EW, padx=5)
        year_combo.set(str(current_year))  # Default to current year
        year_combo.bind('<<ComboboxSelected>>', self.on_selection_change)

        # Refresh Button
        ttk.Button(controls_frame, text="Refresh", command=self.update_summary).grid(row=0, column=4, padx=(20, 5))

        # --- Results Frame ---
        results_frame = ttk.Frame(summary_frame)
        results_frame.pack(fill=tk.X, pady=(10, 0))
        results_frame.columnconfigure(0, weight=1)

        self.total_label = ttk.Label(results_frame, text="Total: ₹0.00", style="Total.TLabel", anchor=tk.CENTER)
        self.total_label.grid(row=0, column=0, pady=5)

        self.paid_label = ttk.Label(results_frame, text="Paid: ₹0.00", style="Paid.TLabel", anchor=tk.CENTER)
        self.paid_label.grid(row=1, column=0, pady=5)

        self.unpaid_label = ttk.Label(results_frame, text="Unpaid: ₹0.00", style="Unpaid.TLabel", anchor=tk.CENTER)
        self.unpaid_label.grid(row=2, column=0, pady=5)

        # --- Company Breakdown Table ---
        company_frame = ttk.LabelFrame(self, text="Top Companies by Month", padding=10)
        company_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=(0, 20))

        cols = ("company", "invoices", "paid", "unpaid", "total")
        self.company_tree = ttk.Treeview(company_frame, columns=cols, show="headings", height=6)

        self.company_tree.heading("company", text="Company")
        self.company_tree.heading("invoices", text="No. of Invoices")
        self.company_tree.heading("paid", text="Paid Amount")
        self.company_tree.heading("unpaid", text="Unpaid Amount")
        self.company_tree.heading("total", text="Total")

        self.company_tree.column("company", width=200, anchor=tk.W)
        self.company_tree.column("invoices", width=100, anchor=tk.CENTER)
        self.company_tree.column("paid", width=120, anchor=tk.E)
        self.company_tree.column("unpaid", width=120, anchor=tk.E)
        self.company_tree.column("total", width=120, anchor=tk.E)

        ysb = ttk.Scrollbar(company_frame, orient="vertical", command=self.company_tree.yview)
        self.company_tree.configure(yscrollcommand=ysb.set)

        ysb.pack(side=tk.RIGHT, fill=tk.Y)
        self.company_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.company_tree.tag_configure('others', font=('Helvetica', 9, 'bold'))

    def on_selection_change(self, event=None):
        self.update_summary()

    def update_summary(self):
        selected_month_name = self.month_var.get()
        selected_year = self.year_var.get()

        if not selected_month_name or not selected_year: return

        month_num = self.month_map[selected_month_name]
        summary = self.db.get_accounting_summary(month_num, selected_year)

        total, paid, unpaid = summary or (None, None, None)
        total = total or 0.0
        paid = paid or 0.0
        unpaid = unpaid or 0.0

        self.total_label.config(text=f"Total: {format_currency(total)}")
        self.paid_label.config(text=f"Paid: {format_currency(paid)}")
        self.unpaid_label.config(text=f"Unpaid: {format_currency(unpaid)}")

        company_data = self.db.get_company_summary_by_month(month_num, selected_year)
        self._update_company_table(company_data)

    def _update_company_table(self, company_data):
        for i in self.company_tree.get_children():
            self.company_tree.delete(i)

        if not company_data: return

        others_invoices = 0
        others_paid = 0.0
        others_unpaid = 0.0
        others_total = 0.0

        for i, row in enumerate(company_data):
            name = row[0]
            invoices = row[1] or 0
            paid = row[2] or 0.0
            unpaid = row[3] or 0.0
            total = row[4] or 0.0

            if i < 5:
                formatted_row = (name, invoices, format_currency(paid), format_currency(unpaid), format_currency(total))
                self.company_tree.insert("", tk.END, values=formatted_row)
            else:
                others_invoices += invoices
                others_paid += paid
                others_unpaid += unpaid
                others_total += total

        if others_total > 0:
            others_row = ("Others", others_invoices, format_currency(others_paid), format_currency(others_unpaid),
                          format_currency(others_total))
            self.company_tree.insert("", tk.END, values=others_row, tags=('others',))
# --- END: Dashboard Page ---
# --- NEW: Sales Commission Page ---
class SalesCommissionPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent); self.db = db; self.app = app
        self.vars = {}
        self.commission_items = []
        self.editing_commission_id = None
        self._setup_widgets()
        self.clear_form()

    def _setup_widgets(self):
        # --- Main Details Form ---
        form = ttk.LabelFrame(self, text="Commission Details", padding=10)
        form.pack(fill=tk.X, padx=10, pady=5)
        form.columnconfigure(1, weight=1); form.columnconfigure(3, weight=1); form.columnconfigure(5, weight=1)

        fields = ["Commission No", "Mechanic Name", "Mobile Number", "Invoice No", "Date of Issue", "Status"]
        self.vars['commission_no'] = tk.StringVar()
        self.vars['mechanic_name'] = tk.StringVar()
        self.vars['mobile_number'] = tk.StringVar()
        self.vars['invoice_no'] = tk.StringVar()
        self.vars['issue_date'] = tk.StringVar()
        self.vars['status'] = tk.StringVar()

        # Row 0
        ttk.Label(form, text="Commission No:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(form, textvariable=self.vars['commission_no'], state='readonly').grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)
        
        ttk.Label(form, text="Mechanic Name:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(form, textvariable=self.vars['mechanic_name']).grid(row=0, column=3, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(form, text="Mobile Number:").grid(row=0, column=4, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(form, textvariable=self.vars['mobile_number']).grid(row=0, column=5, sticky=tk.EW, padx=5, pady=2)

        # Row 1
        ttk.Label(form, text="Invoice No:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(form, textvariable=self.vars['invoice_no']).grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(form, text="Date of Issue:").grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)
        if CALENDAR_ENABLED:
            DateEntry(form, textvariable=self.vars['issue_date'], date_pattern='dd-mm-yyyy').grid(row=1, column=3, sticky=tk.EW, padx=5, pady=2)
        else:
            ttk.Entry(form, textvariable=self.vars['issue_date']).grid(row=1, column=3, sticky=tk.EW, padx=5, pady=2)
            ttk.Label(form, text="(dd-mm-yyyy)").grid(row=1, column=4, sticky=tk.W, padx=5, pady=2)

        ttk.Label(form, text="Status:").grid(row=1, column=4, sticky=tk.W, padx=5, pady=2)
        ttk.Combobox(form, textvariable=self.vars['status'], values=["Pending", "Paid"], state='readonly').grid(row=1, column=5, sticky=tk.EW, padx=5, pady=2)
        
        # --- Item Entry Form ---
        item_form = ttk.LabelFrame(self, text="Add Commission Item", padding=10)
        item_form.pack(fill=tk.X, padx=10, pady=5)
        item_form.columnconfigure(1, weight=3); item_form.columnconfigure(3, weight=1)
        item_form.columnconfigure(5, weight=1); item_form.columnconfigure(7, weight=1)

        self.item_vars = {
            'description': tk.StringVar(),
            'base_amount': tk.DoubleVar(value=0.0),
            'percentage': tk.StringVar(),
            'rupees': tk.StringVar()
        }

        # --- Define Entry Widgets ---
        ttk.Label(item_form, text="Description:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        desc_entry = ttk.Entry(item_form, textvariable=self.item_vars['description'])
        desc_entry.grid(row=0, column=1, sticky=tk.EW, padx=5)

        ttk.Label(item_form, text="Base Amount:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        base_entry = ttk.Entry(item_form, textvariable=self.item_vars['base_amount'])
        base_entry.grid(row=0, column=3, sticky=tk.EW, padx=5)

        ttk.Label(item_form, text="Percentage (%):").grid(row=0, column=4, sticky=tk.W, padx=5, pady=2)
        perc_entry = ttk.Entry(item_form, textvariable=self.item_vars['percentage'])
        perc_entry.grid(row=0, column=5, sticky=tk.EW, padx=5)

        ttk.Label(item_form, text="Rupees (₹):").grid(row=0, column=6, sticky=tk.W, padx=5, pady=2)
        rs_entry = ttk.Entry(item_form, textvariable=self.item_vars['rupees'])
        rs_entry.grid(row=0, column=7, sticky=tk.EW, padx=5)

        ttk.Button(item_form, text="Add Item", command=self.add_item).grid(row=0, column=8, sticky=tk.E, padx=10, pady=2)

        # --- ADD THESE BINDINGS ---
        on_enter_key = lambda event: self.add_item()
        desc_entry.bind("<Return>", on_enter_key)
        base_entry.bind("<Return>", on_enter_key)
        perc_entry.bind("<Return>", on_enter_key)
        rs_entry.bind("<Return>", on_enter_key)

        # --- Items Table ---
        table_frame = ttk.LabelFrame(self, text="Commission Items", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ("sn", "description", "base_amount", "percentage", "rupees", "commission_amount")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
        
        self.tree.heading("sn", text="SN"); self.tree.column("sn", width=40, anchor=tk.CENTER)
        self.tree.heading("description", text="Description"); self.tree.column("description", width=300, anchor=tk.W)
        self.tree.heading("base_amount", text="Base Amount"); self.tree.column("base_amount", width=100, anchor=tk.E)
        self.tree.heading("percentage", text="Perc. %"); self.tree.column("percentage", width=80, anchor=tk.E)
        self.tree.heading("rupees", text="Rupees ₹"); self.tree.column("rupees", width=80, anchor=tk.E)
        self.tree.heading("commission_amount", text="Comm. Amount"); self.tree.column("commission_amount", width=120, anchor=tk.E)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # --- Footer Controls ---
        footer = ttk.Frame(self, padding=10); footer.pack(fill=tk.X)
        
        self.total_label = ttk.Label(footer, text="Total Commission: ₹0.00", font=("Helvetica", 12, "bold"))
        self.total_label.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(footer, text="Clear All", command=self.clear_form).pack(side=tk.RIGHT, padx=5)
        if PDF_ENABLED:
            ttk.Button(footer, text="Save & Download Slip", command=lambda: self.save_commission(download=True)).pack(side=tk.RIGHT, padx=5)
        ttk.Button(footer, text="Save Commission", command=lambda: self.save_commission(download=False)).pack(side=tk.RIGHT, padx=5)
        ttk.Button(footer, text="Delete Selected Item", command=self.delete_item).pack(side=tk.RIGHT, padx=5)

    # --- ADDED METHOD ---
    def add_item(self):
        desc = self.item_vars['description'].get()
        if not desc:
            messagebox.showwarning("Input Error", "Description cannot be empty.")
            return

        try:
            base_amount = self.item_vars['base_amount'].get()
        except tk.TclError:
            messagebox.showwarning("Input Error", "Base Amount must be a valid number.")
            return
            
        perc_str = self.item_vars['percentage'].get()
        rs_str = self.item_vars['rupees'].get()

        perc = None
        rs = None
        commission_amount = 0.0

        try:
            if perc_str not in (None, ""):
                perc = float(perc_str)
            if rs_str not in (None, ""):
                rs = float(rs_str)
        except (ValueError, tk.TclError):
            messagebox.showwarning("Input Error", "Percentage and Rupees must be numbers.")
            return

        if perc is not None and rs is not None:
            messagebox.showwarning("Input Error", "Please enter Percentage OR Rupees, not both.")
            return

        if perc is not None:
            commission_amount = base_amount * (perc / 100.0)
        elif rs is not None:
            commission_amount = rs
        else:
            commission_amount = 0.0 # No commission if no rule

        item_data = {
            'description': desc,
            'base_amount': base_amount,
            'percentage': perc,
            'rupees': rs,
            'final_amount': commission_amount
        }
        self.commission_items.append(item_data)
        self.refresh_items_tree()
        self.clear_item_fields()
    # --- END ADDED METHOD ---

    def delete_item(self):
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select an item to delete.")
            return
        
        selected_iid = self.tree.selection()[0]
        selected_index = self.tree.index(selected_iid)
        
        try:
            del self.commission_items[selected_index]
            self.refresh_items_tree()
        except IndexError:
            messagebox.showerror("Error", "Could not delete the selected item.")

    def refresh_items_tree(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        total_commission = 0.0
        for i, item in enumerate(self.commission_items):
            self.tree.insert("", tk.END, iid=i, values=(
                i + 1,
                item['description'],
                format_currency(item['base_amount']),
                f"{item['percentage']:.2f}%" if item['percentage'] is not None else "",
                format_currency(item['rupees']) if item['rupees'] is not None else "",
                format_currency(item['final_amount'])
            ))
            total_commission += item['final_amount']
            
        self.total_label.config(text=f"Total Commission: {format_currency(total_commission)}")

    def clear_item_fields(self):
        self.item_vars['description'].set("")
        self.item_vars['base_amount'].set(0.0)
        self.item_vars['percentage'].set("")
        self.item_vars['rupees'].set("")
        
    def clear_form(self):
        self.editing_commission_id = None
        self.vars['commission_no'].set(self.db.generate_commission_no())
        self.vars['mechanic_name'].set("")
        self.vars['mobile_number'].set("")
        self.vars['invoice_no'].set("")
        self.vars['issue_date'].set(datetime.now().strftime('%d-%m-%Y'))
        self.vars['status'].set("Pending")
        
        self.commission_items.clear()
        self.refresh_items_tree()
        self.clear_item_fields()
        
    def save_commission(self, download=False):
        mech_name = self.vars['mechanic_name'].get()
        if not mech_name:
            messagebox.showwarning("Input Error", "Mechanic Name is required.")
            return
            
        if not self.commission_items:
            messagebox.showwarning("Input Error", "Cannot save an empty commission slip.")
            return

        total_commission = sum(item['final_amount'] for item in self.commission_items)
        
        main_data = {
            'mechanic_name': mech_name,
            'mobile_number': self.vars['mobile_number'].get(),
            'invoice_no': self.vars['invoice_no'].get(),
            'issue_date': self.vars['issue_date'].get(),
            'status': self.vars['status'].get(),
            'total_amount': total_commission
        }
        
        try:
            commission_id_to_download = None
            if self.editing_commission_id:
                # --- This is an UPDATE ---
                self.db.update_commission(self.editing_commission_id, main_data, self.commission_items)
                commission_id_to_download = self.editing_commission_id
                messagebox.showinfo("Success", f"Commission Slip {self.vars['commission_no'].get()} updated successfully.")
            else:
                # --- This is a CREATE NEW ---
                main_data['commission_no'] = self.vars['commission_no'].get() # Add commission_no for create
                commission_id = self.db.save_commission(main_data, self.commission_items)
                commission_id_to_download = commission_id
                messagebox.showinfo("Success", f"Commission Slip {main_data['commission_no']} saved successfully.")
            
            if download and PDF_ENABLED and commission_id_to_download:
                self.download_slip(commission_id_to_download)

            self.clear_form()
            self.app.commission_history_page.refresh_history()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to save commission: {e}")

    def load_commission_for_edit(self, commission_id):
        """Loads an existing commission slip into the form for editing."""
        self.clear_form() # Reset form and set editing_id to None
        main_details, items = self.db.get_commission_details(commission_id)
        
        if not main_details:
            messagebox.showerror("Error", "Could not load commission details.")
            return
            
        # Set editing ID *after* clear_form
        self.editing_commission_id = commission_id 
        
        # Populate main form
        self.vars['commission_no'].set(main_details['commission_no'])
        self.vars['mechanic_name'].set(main_details['mechanic_name'])
        self.vars['mobile_number'].set(main_details['mobile_number'])
        self.vars['invoice_no'].set(main_details['invoice_no'])
        self.vars['issue_date'].set(main_details['issue_date'])
        self.vars['status'].set(main_details['status'])
        
        # Populate items
        self.commission_items = items
        self.refresh_items_tree()


    def download_slip(self, commission_id):
        if not PDF_ENABLED:
            messagebox.showerror("PDF Error", "PDF library 'reportlab' is not installed.")
            return
            
        main_details, items = self.db.get_commission_details(commission_id)
        if not main_details:
            messagebox.showerror("Error", "Could not find commission details to print.")
            return
            
        filename = f"Commission_{main_details['commission_no']}_{main_details['mechanic_name'].replace(' ', '_')}.pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")], initialfile=filename)
        if not filepath: return
        def draw_watermark(canvas, doc):
            """
            This function is called for every page.
            It draws the image on top of the content.
            """
            # --- Configuration ---
            image_path = resource_path("branding.png")  # <--- IMPORTANT: Change this

            # Check if the image file exists
            if not os.path.exists(image_path):
                print(f"WARNING: Watermark image not found at {image_path}")
                return  # Silently skip if no image

            canvas.saveState()

            # --- Sizing (3x3 inch as you requested) ---
            img_width = 1 * inch
            img_height = 1 * inch

            # --- Positioning (Top-Right) ---
            # Get page size (A4)
            page_width, page_height = A4

            # Set a 0.5-inch margin from the top-right edges
            margin = 0.5 * inch
            x = page_width - img_width - margin
            y = page_height - img_height - margin

            # --- Draw the Image ---
            #canvas.setFillAlpha(0.3)  # Set transparency (e.g., 30% opaque)
            canvas.drawImage(
                image_path,
                x,
                y,
                width=img_width,
                height=img_height,
                mask='auto',  # Handles PNG transparency
                preserveAspectRatio=True
            )

            canvas.restoreState()
        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            styles = getSampleStyleSheet()
            elements = []

            title_style = ParagraphStyle('TitleCustom', parent=styles['h1'], alignment=TA_CENTER)
            elements.append(Paragraph("COMMISSION SLIP", title_style))
            elements.append(Spacer(1, 0.2*inch))

            header_data = [
                ['Commission No:', escape(main_details['commission_no']), 'Date:', escape(main_details['issue_date'])],
                ['Mechanic Name:', escape(main_details['mechanic_name']), 'Status:', escape(main_details['status'])],
                ['Mobile Number:', escape(main_details['mobile_number']), 'Invoice No:', escape(main_details['invoice_no'])],
            ]
            header_table = Table(header_data, colWidths=[1.2*inch, 2.8*inch, 1.2*inch, 2.3*inch])
            header_table.setStyle(TableStyle([
                ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
                ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.2*inch))

            # Table Data
            data = [["SN", "Description", "Base Amount", "Percentage", "Rupees", "Commission"]]
            
            for i, item in enumerate(items, 1):
                data.append([
                    i,
                    Paragraph(escape(item['description']), styles['Normal']),
                    format_currency(item['base_amount']),
                    f"{item['percentage']:.2f}%" if item['percentage'] is not None else "",
                    format_currency(item['rupees']) if item['rupees'] is not None else "",
                    format_currency(item['final_amount'])
                ])
                
            table = Table(data, colWidths=[0.4*inch, 3.1*inch, 1.1*inch, 1*inch, 1*inch, 1.2*inch], repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')), ('TEXTCOLOR',(0,0),(-1,0), colors.whitesmoke),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (1,1), (1,-1), 'LEFT'), # Description column
                ('ALIGN', (2,1), (-1,-1), 'RIGHT'), # All amount columns
                ('VALIGN', (1,1), (1,-1), 'TOP'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.2*inch))
            
            # Total
            total_style = ParagraphStyle('TotalStyle', parent=styles['h3'], alignment=TA_RIGHT)
            elements.append(Paragraph(f"Total Commission: {format_currency(main_details['total_amount'])}", total_style))
            
            doc.build(elements, onFirstPage=draw_watermark, onLaterPages=draw_watermark)
            messagebox.showinfo("Success", f"Commission slip exported to\n{filepath}")
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"An error occurred: {e}")

# --- NEW: Commission History Page ---
# ... (CommissionHistoryPage class remains the same) ...
class CommissionHistoryPage(ttk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent); self.db = db; self.app = app
        self._setup_widgets()
        
    def _setup_widgets(self):
        controls = ttk.Frame(self, padding=10); controls.pack(fill=tk.X)
        ttk.Button(controls, text="Refresh", command=lambda: self.refresh_history(None)).pack(side=tk.LEFT)
        if PDF_ENABLED:
            ttk.Button(controls, text="Download Slip", command=self.download_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(controls, text="Edit Selected", command=self.edit_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(controls, text="Delete Selected", command=self.delete_selected).pack(side=tk.LEFT, padx=5)
        
        # --- NEW SEARCH FRAME ---
        search_frame = ttk.Frame(controls)
        search_frame.pack(side=tk.RIGHT, padx=5)
        
        ttk.Label(search_frame, text="Search by Comm. No:").pack(side=tk.LEFT)
        self.comm_no_search_var = tk.StringVar()
        self.comm_no_entry = ttk.Entry(search_frame, textvariable=self.comm_no_search_var, width=15)
        self.comm_no_entry.pack(side=tk.LEFT, padx=(5,2))

        ttk.Label(search_frame, text="or Mobile No:").pack(side=tk.LEFT, padx=(10, 0))
        self.mobile_search_var = tk.StringVar()
        self.mobile_entry = ttk.Entry(search_frame, textvariable=self.mobile_search_var, width=15)
        self.mobile_entry.pack(side=tk.LEFT, padx=(5,2))
        
        ttk.Button(search_frame, text="Search", command=self.search_commissions).pack(side=tk.LEFT, padx=(2,0))
        
        # Bindings to clear other field
        self.comm_no_entry.bind("<KeyRelease>", lambda e: self.mobile_search_var.set(""))
        self.mobile_entry.bind("<KeyRelease>", lambda e: self.comm_no_search_var.set(""))
        # --- END NEW SEARCH FRAME ---

        table_frame = ttk.LabelFrame(self, text="Commission History", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ("id", "commission_no", "mechanic_name", "invoice_no", "issue_date", "status", "total_amount")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        
        self.tree.column("id", width=0, stretch=tk.NO)
        self.tree.heading("commission_no", text="Comm. No"); self.tree.column("commission_no", width=80, anchor=tk.W)
        self.tree.heading("mechanic_name", text="Mechanic Name"); self.tree.column("mechanic_name", width=200, anchor=tk.W)
        self.tree.heading("invoice_no", text="Invoice No"); self.tree.column("invoice_no", width=100, anchor=tk.W)
        self.tree.heading("issue_date", text="Date"); self.tree.column("issue_date", width=100, anchor=tk.CENTER)
        self.tree.heading("status", text="Status"); self.tree.column("status", width=80, anchor=tk.W)
        self.tree.heading("total_amount", text="Total Amount"); self.tree.column("total_amount", width=120, anchor=tk.E)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.refresh_history()
        
    def refresh_history(self, data=None):
        # Clear search bars if refreshing all
        if data is None:
            if hasattr(self, 'comm_no_search_var'):
                self.comm_no_search_var.set("")
            if hasattr(self, 'mobile_search_var'):
                self.mobile_search_var.set("")

        for i in self.tree.get_children():
            self.tree.delete(i)
        
        if data is None:
            data = self.db.get_all_commissions()
        
        for row in data:
            formatted = list(row)
            formatted[6] = format_currency(row[6]) # Format total amount
            self.tree.insert("", tk.END, values=formatted)
            
    def download_selected(self):
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a commission slip to download.")
            return
            
        selected_iid = self.tree.selection()[0]
        commission_id = self.tree.item(selected_iid, 'values')[0]
        
        # Re-use the download logic from the main commission page
        self.app.commission_page.download_slip(commission_id)

    def search_commissions(self):
        """Filters the commission list based on search criteria."""
        comm_no = self.comm_no_search_var.get().strip()
        mobile = self.mobile_search_var.get().strip()
        
        if not comm_no and not mobile:
            self.refresh_history(None) # Refresh all if both are empty
            return
            
        results = self.db.search_commissions(commission_no=comm_no, mobile_number=mobile)
        
        if not results:
            messagebox.showinfo("No Results", "No commission slips found matching that criteria.")
        
        self.refresh_history(results)

    def edit_selected(self):
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a commission slip to edit.")
            return
            
        selected_iid = self.tree.selection()[0]
        commission_id = self.tree.item(selected_iid, 'values')[0]
        
        # Call the new load method on the commission page
        self.app.commission_page.load_commission_for_edit(commission_id)
        
        # Switch to the commission page
        self.app.notebook.select(self.app.commission_page)

    def delete_selected(self):
        if not self.tree.selection():
            messagebox.showwarning("No Selection", "Please select a commission slip to delete.")
            return
            
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to permanently delete this commission record?"):
            return
            
        try:
            selected_iid = self.tree.selection()[0]
            commission_id = self.tree.item(selected_iid, 'values')[0]
            self.db.delete_commission(commission_id)
            self.refresh_history()
            messagebox.showinfo("Success", "Commission record deleted.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete record: {e}")


class AboutPage(ttk.Frame):
    def __init__(self, parent, db, app, db_path):
        super().__init__(parent)
        self.db = db
        self.app = app
        self.db_path = db_path
        self._setup_widgets()

    def _setup_widgets(self):
        # Main container with padding, using pack for a clean vertical layout
        main_frame = ttk.Frame(self, padding=(30, 20))
        main_frame.pack(expand=True, fill=tk.BOTH)

        # --- Title ---
        title_label = ttk.Label(main_frame, text="Kabir Auto Parts - Shop Management System",
                                font=("Helvetica", 24, "bold"), anchor="center")
        title_label.pack(fill=tk.X, pady=(10, 5))

        # --- Subtitle ---
        desc_label = ttk.Label(
            main_frame,
            text="Your complete solution for Order Management, Invoicing, and Part Tracking.",
            font=("Helvetica", 14),
            anchor="center"
        )
        desc_label.pack(fill=tk.X, pady=(0, 20))

        # --- Separator ---
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)

        # --- About the Project ---
        dev_frame = ttk.LabelFrame(main_frame, text="About this Project", padding=15)
        dev_frame.pack(fill=tk.X, pady=10, padx=20)

        # Developer Info
        dev_text_1 = "This application was developed by:"
        dev_label_1 = ttk.Label(dev_frame, text=dev_text_1, font=("Helvetica", 11), anchor="center")
        dev_label_1.pack(fill=tk.X)

        dev_text_2 = "ash4code (Ash)"
        dev_label_2 = ttk.Label(dev_frame, text=dev_text_2, font=("Helvetica", 12, "bold"), anchor="center")
        dev_label_2.pack(fill=tk.X, pady=(2, 5))

        # Purpose
        purpose_text = "It was custom-built as a comprehensive management solution for Kabir Auto Parts."
        purpose_label = ttk.Label(dev_frame, text=purpose_text, font=("Helvetica", 11), anchor="center",
                                  justify=tk.CENTER, wraplength=700)
        purpose_label.pack(fill=tk.X, pady=5)

        # AI Acknowledgment
        ai_text = ("This project represents a modern development workflow. I (a generative AI) "
                   "functioned as a pair-programming partner, collaborating with the developer (ash4code) "
                   "to help design, write, and debug the application.")
        ai_label = ttk.Label(dev_frame, text=ai_text, font=("Helvetica", 10, "italic"), anchor="center",
                             justify=tk.CENTER, wraplength=700)
        ai_label.pack(fill=tk.X, pady=(10, 5))

        # --- Support & Contact ---
        support_frame = ttk.LabelFrame(main_frame, text="Support & Contact", padding=15)
        support_frame.pack(fill=tk.X, pady=10, padx=20)

        ttk.Label(support_frame, text="Developer: Ash", font=("Helvetica", 11), anchor="center").pack(fill=tk.X)
        ttk.Label(support_frame, text="GitHub: github.com/ash4code", font=("Helvetica", 11), anchor="center").pack(
            fill=tk.X)
        ttk.Label(support_frame, text="Email: ashkhan2101@gmail.com", font=("Helvetica", 11), anchor="center").pack(
            fill=tk.X)

        # --- Database Management (FIXED) ---
        db_frame = ttk.LabelFrame(main_frame, text="Database Management", padding=15)
        db_frame.pack(fill=tk.X, pady=10, padx=20)

        # Use a new frame to hold the buttons side-by-side
        db_btn_frame = ttk.Frame(db_frame)
        db_btn_frame.pack(fill=tk.X, expand=True)

        # Use pack(side=LEFT) to put them next to each other
        # The fill=X and expand=True will make them share the space
        ttk.Button(db_btn_frame, text="Export Database Backup", command=self.export_database).pack(side=tk.LEFT,
                                                                                                   fill=tk.X,
                                                                                                   expand=True,
                                                                                                   padx=(0, 5))
        ttk.Button(db_btn_frame, text="Import Database Backup", command=self.import_database).pack(side=tk.LEFT,
                                                                                                   fill=tk.X,
                                                                                                   expand=True,
                                                                                                   padx=(5, 0))

        warning_label = ttk.Label(
            db_frame,
            text="WARNING: Importing a database will overwrite all current data and restart the application.",
            font=("Helvetica", 9, "italic"),
            foreground="red",
            anchor="center"
        )
        warning_label.pack(fill=tk.X, pady=(10, 0))

    def export_database(self):
        """Saves a copy of the current database file."""
        try:
            now = datetime.now().strftime("%Y-%m-%d_%H%M")
            filename = f"pos_backup_{now}.db"

            filepath = filedialog.asksaveasfilename(
                defaultextension=".db",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")],
                initialfile=filename
            )

            if not filepath:
                return  # User cancelled

            shutil.copyfile(self.db_path, filepath)
            messagebox.showinfo("Export Success", f"Database successfully exported to:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred during export:\n{e}")

    def import_database(self):
        """Overwrites the current database with an imported one and restarts."""

        # Critical Warning
        if not messagebox.askyesno("Confirm Import",
                                   "WARNING: This will overwrite the current database and restart the application.\n\n"
                                   "All unsaved data will be lost.\n\n"
                                   "Are you sure you want to continue?"):
            return

        try:
            filepath = filedialog.askopenfilename(
                title="Select Database Backup to Import",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")]
            )

            if not filepath:
                return  # User cancelled

            # --- CRITICAL STEPS ---
            # 1. Close the current database connection to release the file lock
            self.db.conn.close()

            # 2. Overwrite the database file
            shutil.copyfile(filepath, self.db_path)

            # 3. Inform user and restart
            messagebox.showinfo("Import Success", "Database imported successfully.\nThe application will now restart.")

            # 4. Restart the application
            os.execl(sys.executable, sys.executable, *sys.argv)

        except Exception as e:
            messagebox.showerror("Import Error",
                                 f"An error occurred during import:\n{e}\n\nThe application might be in an unstable state. Please restart it.")


class App(tk.Tk):
    def __init__(self, db, db_path):
        super().__init__()
        self.db = db
        self.db_path = db_path
        self.version = "v1.1.0"
        self.after_idle(self._configure_window)
        self.PASSWORD = "admin123"
        self.accounting_unlocked = False
        self.commission_unlocked = False

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        self.notebook.bind("<Button-1>", self.on_tab_click, True)

        # Style configuration
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.map('Treeview', background=[('selected', '#0078D7')], foreground=[('selected', '#FFFFFF')])

        # --- INSTANTIATE PAGES ---
        self.dashboard_page = DashboardPage(self.notebook, self.db, self)

        # 1. The NEW Hero Page (Complex Mode)
        self.hero_order_page = HeroOrderPage(self.notebook, self.db, self)

        # 2. The STANDARD Order Page (Simple Mode)
        self.order_page = OrderPage(self.notebook, self.db, self)

        self.history_page = HistoryPage(self.notebook, self.db, self)
        self.nonoem_page = NonOEMOrderPage(self.notebook, self.db, self)
        self.part_request_page = PartRequestPage(self.notebook, self.db, self)
        self.commission_page = SalesCommissionPage(self.notebook, self.db, self)
        self.commission_history_page = CommissionHistoryPage(self.notebook, self.db, self)
        self.accounting_page = AccountingPage(self.notebook, self.db, self)
        self.about_page = AboutPage(self.notebook, self.db, self, self.db_path)

        # --- ADD TABS TO NOTEBOOK ---
        self.notebook.add(self.dashboard_page, text="Dashboard")
        self.notebook.add(self.hero_order_page, text="Hero Order")
        self.notebook.add(self.order_page, text="Create Order")
        self.notebook.add(self.history_page, text="Order History")
        self.notebook.add(self.nonoem_page, text="Non-OEM Order")
        self.notebook.add(self.part_request_page, text="Part Request")
        self.notebook.add(self.commission_page, text="Sales Commission")
        self.notebook.add(self.commission_history_page, text="Commission History")
        self.notebook.add(self.accounting_page, text="Accounting")
        self.notebook.add(self.about_page, text="About")

    def on_tab_click(self, event):
        try:
            clicked_tab_index = self.notebook.index(f"@{event.x},{event.y}")
            tab_text = self.notebook.tab(clicked_tab_index, "text")
        except tk.TclError:
            return

        if tab_text == "Dashboard":
            try:
                self.dashboard_page.update_summary()
            except:
                pass

        if tab_text == "Accounting" and not self.accounting_unlocked:
            current_tab = self.notebook.select()
            password = simpledialog.askstring("Password", "Enter admin password:", show='*')
            if password == self.PASSWORD:
                self.accounting_unlocked = True
            else:
                if password is not None: messagebox.showerror("Denied", "Incorrect password.")
                self.notebook.select(current_tab);
                return "break"

        elif tab_text in ["Sales Commission", "Commission History"] and not self.commission_unlocked:
            current_tab = self.notebook.select()
            password = simpledialog.askstring("Password", "Enter admin password:", show='*')
            if password == self.PASSWORD:
                self.commission_unlocked = True
            else:
                if password is not None: messagebox.showerror("Denied", "Incorrect password.")
                self.notebook.select(current_tab);
                return "break"

    def _configure_window(self):
        self.title(f"POS for Kabir Auto Parts - {self.version}")
        self.geometry("1200x800")
        try:
            icon_path = resource_path("icon.png")
            self.icon = tk.PhotoImage(file=icon_path)
            self.iconphoto(True, self.icon)
        except:
            pass

if __name__ == "__main__":
    # --- Define a persistent location for the database ---
    APP_NAME = "KabirAutoPOS"
    persistent_folder = os.path.join(os.getenv('APPDATA'), APP_NAME)
    os.makedirs(persistent_folder, exist_ok=True)
    persistent_db_path = os.path.join(persistent_folder, "database.db")

    # --- Check if the persistent database exists ---
    if not os.path.exists(persistent_db_path):
        try:
            # Find the path to the bundled database (using resource_path)
            bundled_db_path = resource_path("database.db")

            # --- NEW CHECK: Only copy if the file actually exists ---
            if os.path.exists(bundled_db_path):
                shutil.copyfile(bundled_db_path, persistent_db_path)
                print(f"Database copied to: {persistent_db_path}")
            else:
                print("No bundled database found. A fresh, empty database will be created automatically.")

        except Exception as e:
            # If copying fails, just print a warning and continue. Don't crash.
            print(f"Warning: Could not copy bundled DB ({e}). Continuing with fresh DB.")

    threading.Thread(target=init_browser_thread, daemon=True).start()

    # --- Always use the persistent database path ---
    try:
        db_instance = Database(persistent_db_path)
        app = App(db_instance, persistent_db_path)
        app.mainloop()
    except Exception as e:
        root = tk.Tk();
        root.withdraw()
        messagebox.showerror("Application Error", f"An unexpected error occurred:\n{e}")
        sys.exit(1)
